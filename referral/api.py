import frappe
from collections import defaultdict
from deep_translator import GoogleTranslator
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate
from frappe.utils import getdate, today

VALID_DEPARTMENTS = [
    "Medicine", "Gynaecology", "Orthopedics", "Spine",
    "Surgery", "Dental", "Mental Health Clinic", "Rheumatology OPD",
    "Cardiology", "Dermatology", "Diabetology", "ENT",
    "Gastrology", "Head & Neck", "Neurology + Epilepsy", "Oncology",
    "Pulmonology", "Sickle Cell", "Cataract Surgery", "Ophthalmology",
    "Plastic Surgery", "Urology", "Pain Management OPD", "Physiotherapy",
    "Others", "Other"
]

REGULAR_OPD_DEPTS = ["Medicine", "Gynaecology", "Orthopedics", "Spine", "Surgery", "Dental", "Mental Health Clinic", "Pain Management OPD", "Rheumatology OPD", "Physiotherapy", "Others", "Other"]
SPECIALIST_OPD_DEPTS = ["Cardiology", "Dermatology", "Diabetology", "ENT", "Gastrology", "Head & Neck", "Neurology + Epilepsy", "Oncology", "Pulmonology", "Sickle Cell", "Plastic Surgery", "Urology", "Pain Management OPD", "Rheumatology OPD", "Others", "Other"]
SURGICAL_OPD_DEPTS = ["Cataract Surgery", "Ophthalmology", "Plastic Surgery", "Urology", "Pain Management OPD", "Others", "Other"]

# Hindi/Marathi keywords → English OPD department mapping
DEPARTMENT_KEYWORDS = {
    # Backward-compatible English aliases
    "Epilepsy": "Neurology + Epilepsy",
    "Neurology": "Neurology + Epilepsy",
    "Psychiatry": "Mental Health Clinic",
    "Mental Health Clinic": "Mental Health Clinic",
    "Pain Clinic": "Pain Management OPD",
    "Pain Management": "Pain Management OPD",
    "Pain Management OPD": "Pain Management OPD",
    "Rheumatology": "Rheumatology OPD",
    "Rheumatology OPD": "Rheumatology OPD",
    "Gastroenterology": "Gastrology",
    # Gynaecology
    "स्त्रीरोग": "Gynaecology",
    "स्त्री रोग": "Gynaecology",
    "गायनेकॉलॉजी": "Gynaecology",
    "प्रसूती": "Gynaecology",
    "प्रसूतिशास्त्र": "Gynaecology",
    "प्रसूतिशास्र": "Gynaecology",
    "प्रसूति शास्त्र": "Gynaecology",
    "स्त्रीरोगशास्त्र": "Gynaecology",
    # Orthopedics
    "हाडरोग": "Orthopedics",
    "अस्थिरोग": "Orthopedics",
    "ऑर्थोपेडिक्स": "Orthopedics",
    "हड्डी": "Orthopedics",
    "हड्डी रोग": "Orthopedics",
    # Spine
    "पाठीचा कणा": "Spine",
    "स्पाइन": "Spine",
    "मणका": "Spine",
    "रीढ़": "Spine",
    "रीढ़ की हड्डी": "Spine",
    # Cardiology
    "हृदयरोग": "Cardiology",
    "कार्डिओलॉजी": "Cardiology",
    "हृदय": "Cardiology",
    "दिल": "Cardiology",
    "हृदयरोगशास्त्र": "Cardiology",
    "कार्डियलजी": "Cardiology",
    # Mental Health / Psychiatry
    "मानसिक आरोग्य": "Mental Health Clinic",
    "मानसिक": "Mental Health Clinic",
    "मनोविकार": "Mental Health Clinic",
    "मानसोपचार": "Mental Health Clinic",
    "मनोचिकित्सा": "Mental Health Clinic",
    "मानसिक आरोग्य (२ दिवस)": "Mental Health Clinic",
    "मानसिक आरोग्य (2 दिवस)": "Mental Health Clinic",
    "मानसिक स्वास्थ्य (2 दिन)": "Mental Health Clinic",
    "मानसिक स्वास्थ्य (२ दिन)": "Mental Health Clinic",
    "मानसिक स्वास्थ्य": "Mental Health Clinic",
    # Surgery
    "सर्जन": "Surgery",
    "शल्यचिकित्सक": "Surgery",
    "सर्जरी": "Surgery",
    "जनरल सर्जन": "Surgery",
    "शस्त्रक्रिया": "Surgery",
    "शल्य चिकित्सा": "Surgery",
    # Cataract Surgery
    "मोतीबिंदू": "Cataract Surgery",
    "मोतियाबिंद": "Cataract Surgery",
    "मोतीबिंदू शस्त्रक्रिया": "Cataract Surgery",
    "मोतियाबिंद सर्जरी": "Cataract Surgery",
    # Ophthalmology
    "नेत्र": "Ophthalmology",
    "डोळा": "Ophthalmology",
    "नेत्रविज्ञान": "Ophthalmology",
    # ENT
    "कान नाक घसा": "ENT",
    "कान": "ENT",
    "ईएनटी": "ENT",
    # Dermatology
    "त्वचा": "Dermatology",
    "त्वचारोग": "Dermatology",
    "चर्मरोग": "Dermatology",
    "त्वचा विज्ञान": "Dermatology",
    # Dental
    "दंत": "Dental",
    "दात": "Dental",
    "दंतचिकित्सा": "Dental",
    "दंतचिकित्सा (३ दिवस)": "Dental",
    "दंतचिकित्सा (3 दिवस)": "Dental",
    "दंत चिकित्सा (3 दिन)": "Dental",
    "दंत चिकित्सा (३ दिन)": "Dental",
    "दंत चिकित्सा": "Dental",
    # Oncology
    "कर्करोग": "Oncology",
    "ऑन्कोलॉजी": "Oncology",
    "कैंसर विज्ञान": "Oncology",
    # Sickle Cell
    "सिकल सेल": "Sickle Cell",
    "विकृतिरक्तकोशिका": "Sickle Cell",
    # Diabetology
    "मधुमेह": "Diabetology",
    "डायाबैटोलोजी": "Diabetology",
    "मधुमेहशास्त्र": "Diabetology",
    # Gastrology
    "पोट": "Gastrology",
    "जठर": "Gastrology",
    "पोटरोग": "Gastrology",
    "गॅस्ट्रोएन्टेरोलॉजी": "Gastrology",
    "गैस्ट्रोएंटरोलॉजी": "Gastrology",
    # Pulmonology
    "फुफ्फुस": "Pulmonology",
    "श्वसन": "Pulmonology",
    "फुफ्फुसशास्त्र": "Pulmonology",
    "फुफ्फुसविज्ञान": "Pulmonology",
    # Head & Neck
    "डोके आणि मान": "Head & Neck",
    "सिर और गर्दन": "Head & Neck",
    # Rheumatology
    "संधिवात": "Rheumatology OPD",
    "सांधेदुखी": "Rheumatology OPD",
    "रुमॅटोलॉजी": "Rheumatology OPD",
    "रुमॅटोलॉजी ओपीडी": "Rheumatology OPD",
    "संधिवातीयशास्त्र": "Rheumatology OPD",
    "रुमेटोलॉजी ओपीडी": "Rheumatology OPD",
    # Neurology + Epilepsy
    "अपस्मार": "Neurology + Epilepsy",
    "मिरगी": "Neurology + Epilepsy",
    "फेफरे": "Neurology + Epilepsy",
    "मज्जातंतू": "Neurology + Epilepsy",
    "न्यूरोलॉजी": "Neurology + Epilepsy",
    "तंत्रिका विज्ञान + मिर्ग": "Neurology + Epilepsy",
    "न्यूरोलॉजी + एपिलेप्सी": "Neurology + Epilepsy",
    # Urology
    "मूत्ररोग": "Urology",
    "यूरोलॉजी": "Urology",
    "उरोलोजि": "Urology",
    "मूत्रविज्ञान": "Urology",
    # Plastic Surgery
    "प्लास्टिक सर्जरी": "Plastic Surgery",
    # Pain Management
    "वेदना": "Pain Management OPD",
    "दुखणे": "Pain Management OPD",
    "दर्द": "Pain Management OPD",
    "पेन": "Pain Management OPD",
    "वेदना व्यवस्थापन": "Pain Management OPD",
    "वेदना व्यवस्थापन ओपीडी": "Pain Management OPD",
    "दर्द प्रबंधन": "Pain Management OPD",
    "दर्द प्रबंधन ओपीडी": "Pain Management OPD",
    "Physiotherapy": "Physiotherapy",
    "फिजिओथेरपी": "Physiotherapy",
    "फिजियोथेरेपी": "Physiotherapy",
    # Medicine
    "औषध": "Medicine",
    "मेडिसिन": "Medicine",
    "दवा": "Medicine",
    # Others
    "इतर": "Others",
    "अन्य": "Others",
}


VILLAGE_MATCH_CACHE_TTL_SECONDS = 600

VILLAGE_MESSAGES = {
    "village_empty": {
        "en": "Village name cannot be empty. Please try again.",
        "hi": "गाँव का नाम खाली नहीं हो सकता। कृपया पुनः प्रयास करें।",
        "mr": "गावाचे नाव रिकामे असू शकत नाही. कृपया पुन्हा प्रयत्न करा.",
    },
    "resolved": {
        "en": "Village confirmed: {name}",
        "hi": "गाँव की पुष्टि हुई: {name}",
        "mr": "गाव निश्चित केले: {name}",
    },
    "no_match_in_taluka": {
        "en": "No matching village found under this taluka. Please type the village name again:",
        "hi": "इस तालुका में यह नाम वाला कोई गाँव नहीं मिला। कृपया गाँव का नाम दोबारा टाइप करें:",
        "mr": "तुमच्या तालुक्यात या नावाचे गाव आढळले नाही. कृपया गावाचे नाव पुन्हा एकदा टाईप करा:",
    },
    "did_you_mean": {
        "en": "Do you mean one of these villages? Please reply with the correct option number:",
        "hi": "क्या आपका मतलब इनमें से कोई गाँव है? कृपया सही विकल्प नंबर के साथ जवाब दें:",
        "mr": "मला खालीलपैकी एक गाव वाटते का? कृपया योग्य पर्याय क्रमांक निवडून पाठवा:",
    },
    "none_of_these": {
        "en": "None of these - Type again",
        "hi": "इनमें से कोई नहीं - दोबारा टाइप करें",
        "mr": "यापैकी काहीही नाही (पुन्हा टाईप करा)",
    },
    "session_expired": {
        "en": "Your village list has expired. Please type the village name again.",
        "hi": "आपकी गाँव सूची समाप्त हो गई है। कृपया गाँव का नाम फिर से टाइप करें।",
        "mr": "तुमची गाव यादी कालबाह्य झाली आहे. कृपया गावाचे नाव पुन्हा टाईप करा.",
    },
}


def village_msg(key: str, lang: str, **kwargs) -> str:
    lang = lang if lang in ("en", "hi", "mr") else "en"
    template = VILLAGE_MESSAGES[key][lang]
    return template.format(**kwargs) if kwargs else template


def village_display_name(village_name: str, lang: str) -> str:
    """Single-language display name for a Village Profile record."""
    if lang == "mr":
        v_mr = frappe.db.get_value("Village Profile", village_name, "village_name_marathi")
        return v_mr or village_name
    return village_name


def is_devanagari(text: str) -> bool:
    """Check if text contains Devanagari script"""
    return any("\u0900" <= c <= "\u097F" for c in (text or ""))


def _iast_to_english(iast_text: str) -> str:
    """
    Convert IAST transliteration to common English spellings.
    Handles digraph mappings and Hindi/Marathi schwa deletion.
    Key: schwa deletion runs BEFORE vowel normalization so only
    implicit 'a' (schwa) is removed, not explicit 'ā' (long a).
    """
    import re
    import unicodedata

    # Step 1: Hindi/Marathi schwa deletion on IAST text
    # Remove trailing implicit 'a' (schwa) but NOT 'ā', 'ī', 'ū' etc.
    # In IAST: implicit schwa = 'a', explicit long vowel = 'ā'
    words = iast_text.split()
    cleaned = []
    for w in words:
        if len(w) > 1 and w[-1] == 'a' and w[-2] not in 'aeiouāīūṛ':
            w = w[:-1]
        cleaned.append(w)
    result = ' '.join(cleaned)

    # Convert anusvara 'ṃ' to 'm' before labials (p, b, ph, bh, v, m), otherwise 'n'
    result = re.sub(r"ṃ(?=[pbvm]|ph|bh)", "m", result)
    result = result.replace("ṃ", "n")

    # Step 2: Apply multi-char IAST → English mappings
    replacements = [
        ("kṣ", "ksh"), ("ṣ", "sh"), ("ś", "sh"),
        ("ch", "chh"), ("c", "ch"),   # IAST 'c' = ch sound
        ("jñ", "gya"),
        ("ṭ", "t"), ("ḍ", "d"), ("ṇ", "n"), ("ṅ", "ng"),
        ("ñ", "n"),
        ("ḥ", "h"),
        ("ā", "a"), ("ī", "i"), ("ū", "u"),
        ("ṛ", "ri"),
        ("ai", "ai"), ("au", "au"),
        ("ē", "e"), ("ō", "o"),
    ]
    for old, new in replacements:
        result = result.replace(old, new)

    # Step 3: Remove any remaining diacritics
    normalized = unicodedata.normalize('NFD', result)
    result = ''.join(
        c for c in normalized
        if unicodedata.category(c) != 'Mn'
    )

    return result


def transliterate_to_roman(text: str) -> str:
    """
    Transliterate Devanagari text to Roman English using indic-transliteration.
    This does script-to-script conversion (preserves names like साक्षी → Sakshi)
    instead of translating meaning (which would give 'Witness').
    Falls back to original text if transliteration fails.
    """
    if not text:
        return text

    import re
    # Clean doctor prefixes in both Devanagari and Latin (case-insensitive, longest match first)
    text_clean = re.sub(r'^(डॉक्टर|डाक्टर|डॉ\.|डॉ|डाॅ\.|डाॅ|डा\.|dr\.|dr|daॉ\.|daॉ|daॅ\.|daॅ)\s*', 'Dr. ', text.strip(), flags=re.IGNORECASE)

    if not is_devanagari(text_clean):
        return text_clean

    try:
        iast = transliterate(text_clean, sanscript.DEVANAGARI, sanscript.IAST)
        result = _iast_to_english(iast).strip().title()
        
        # Ensure "Dr. " has correct casing and is not altered by IAST/Title conversion
        result = re.sub(r'^(dr\.|daॉ\.|daॅ\.)\s*', 'Dr. ', result, flags=re.IGNORECASE)
        
        frappe.logger().info(
            f"Transliterated '{text}' → '{result}'"
        )
        return result if result else text_clean

    except Exception as e:
        frappe.log_error(
            f"Transliteration failed for '{text}': {str(e)}",
            "Transliteration Error"
        )
        return text_clean


def translate_to_english(text: str) -> str:
    """
    Translate Devanagari text to English using deep-translator.
    Used for additional notes where we want meaning not transliteration.
    Falls back to original text if translation fails.
    """
    if not text or not is_devanagari(text):
        return text

    try:
        translated = GoogleTranslator(source="auto", target="en").translate(text)
        result = translated.strip() if translated else text
        frappe.logger().info(
            f"Translated '{text}' → '{result}'"
        )
        return result if result else text

    except Exception as e:
        frappe.log_error(
            f"Translation failed for '{text}': {str(e)}",
            "Translation Error"
        )
        return text


def resolve_village(village_raw: str) -> str | None:
    """
    Resolve village name to Village Profile record.
    Tries: original text, Marathi name, transliterated, case-insensitive, fuzzy.
    Returns None if no match found (never falls back to a random village).
    """
    if not village_raw:
        frappe.logger().warning("Village name is empty — leaving unset")
        return None

    village_raw = village_raw.strip()
    frappe.logger().info(f"[resolve_village] Starting search for: '{village_raw}'")

    # Try original text exact match (English or Marathi)
    village = frappe.db.get_value(
        "Village Profile", {"village_name": village_raw}, "name"
    )
    if village:
        frappe.logger().info(f"[resolve_village] Exact English match found: {village}")
        return village

    # Try by Marathi name if input is Devanagari
    if is_devanagari(village_raw):
        village = frappe.db.get_value(
            "Village Profile", {"village_name_marathi": village_raw}, "name"
        )
        if village:
            frappe.logger().info(f"[resolve_village] Exact Marathi match found: {village}")
            return village

    # Try case-insensitive match on English name
    village = frappe.db.sql("""
        SELECT name FROM `tabVillage Profile`
        WHERE LOWER(village_name) = LOWER(%(name)s)
        LIMIT 1
    """, {"name": village_raw}, as_dict=True)
    if village:
        frappe.logger().info(f"[resolve_village] Case-insensitive English match found: {village[0].name}")
        return village[0].name

    # Try transliterated version
    if is_devanagari(village_raw):
        transliterated = transliterate_to_roman(village_raw)
        frappe.logger().info(f"[resolve_village] Transliterated to: '{transliterated}'")
        if transliterated:
            village = frappe.db.get_value(
                "Village Profile", {"village_name": transliterated}, "name"
            )
            if village:
                frappe.logger().info(f"[resolve_village] Transliterated exact match found: {village}")
                return village

            # Case-insensitive on transliterated
            village = frappe.db.sql("""
                SELECT name FROM `tabVillage Profile`
                WHERE LOWER(village_name) = LOWER(%(name)s)
                LIMIT 1
            """, {"name": transliterated}, as_dict=True)
            if village:
                frappe.logger().info(f"[resolve_village] Transliterated case-insensitive match found: {village[0].name}")
                return village[0].name

            # Bidirectional fuzzy: transliterated starts with village OR village starts with transliterated
            # Also require minimum 3 chars to avoid over-matching
            if len(transliterated) >= 3:
                village = frappe.db.sql("""
                    SELECT name, village_name FROM `tabVillage Profile`
                    WHERE LOWER(village_name) LIKE CONCAT(LOWER(%(name)s), '%%')
                       OR LOWER(%(name)s) LIKE CONCAT(LOWER(village_name), '%%')
                    ORDER BY
                        ABS(CHAR_LENGTH(village_name) - CHAR_LENGTH(%(name)s)) ASC
                    LIMIT 1
                """, {"name": transliterated}, as_dict=True)
                if village:
                    frappe.logger().info(
                        f"[resolve_village] Fuzzy match: '{village_raw}' → '{transliterated}' → '{village[0].village_name}'"
                    )
                    return village[0].name

    # No match found — log available villages and return None
    all_villages = frappe.get_all("Village Profile", fields=["name", "village_name"])
    village_names = [v.get("village_name", v.get("name")) for v in all_villages]
    frappe.logger().warning(
        f"No village match found for '{village_raw}' (transliterated: '{transliterate_to_roman(village_raw) if is_devanagari(village_raw) else 'N/A'}') — leaving unset for manual correction. Available villages: {village_names[:10]}"
    )
    return None


def resolve_phc(phc_raw: str) -> str | None:
    """
    Resolve PHC name to PHC record.
    Tries: original text, Marathi name, transliterated, case-insensitive, fuzzy.
    If 'NA' or 'no PHC' variation, returns None.
    If 'Other' variation or no match is found for other PHC name, returns 'Other'.
    """
    if not phc_raw:
        frappe.logger().warning("PHC name is empty — leaving unset")
        return None

    phc_raw = phc_raw.strip()
    phc_lower = phc_raw.lower()
    frappe.logger().info(f"[resolve_phc] Starting search for: '{phc_raw}'")

    # 1. Handle "NA" / "No PHC" responses
    na_keywords = {
        "na", "n/a", "none", "nil", "no", "not applicable", "not available",
        "नाही", "नाही आहे", "एनए", "एन/ए"
    }
    if phc_lower in na_keywords:
        frappe.logger().info(f"[resolve_phc] Input '{phc_raw}' resolved as NA/None")
        return None

    # 2. Handle explicit "Other" / "इतर" responses
    other_keywords = {
        "other", "other phc", "इतर", "इतर पीएचसी", "इतर पी.एच.सी.", "इतर पी. एच. सी."
    }
    if phc_lower in other_keywords:
        frappe.logger().info(f"[resolve_phc] Input '{phc_raw}' resolved as 'Other'")
        return "Other"

    # 3. Try original exact match (English or Marathi)
    phc = frappe.db.get_value("PHC", {"phc_name": phc_raw}, "name")
    if phc:
        frappe.logger().info(f"[resolve_phc] Exact English match found: {phc}")
        return phc

    # Try by Marathi name if input is Devanagari
    if is_devanagari(phc_raw):
        phc = frappe.db.get_value(
            "PHC", {"phc_name_marathi": phc_raw}, "name"
        )
        if phc:
            frappe.logger().info(f"[resolve_phc] Exact Marathi match found: {phc}")
            return phc

    # Try case-insensitive on English name
    phc = frappe.db.sql("""
        SELECT name FROM `tabPHC`
        WHERE LOWER(phc_name) = LOWER(%(name)s)
        LIMIT 1
    """, {"name": phc_raw}, as_dict=True)
    if phc:
        frappe.logger().info(f"[resolve_phc] Case-insensitive English match found: {phc[0].name}")
        return phc[0].name

    # Try transliterated version
    if is_devanagari(phc_raw):
        transliterated = transliterate_to_roman(phc_raw)
        frappe.logger().info(f"[resolve_phc] Transliterated to: '{transliterated}'")
        if transliterated:
            phc = frappe.db.get_value("PHC", {"phc_name": transliterated}, "name")
            if phc:
                frappe.logger().info(f"[resolve_phc] Transliterated exact match found: {phc}")
                return phc

            phc = frappe.db.sql("""
                SELECT name FROM `tabPHC`
                WHERE LOWER(phc_name) = LOWER(%(name)s)
                LIMIT 1
            """, {"name": transliterated}, as_dict=True)
            if phc:
                frappe.logger().info(f"[resolve_phc] Transliterated case-insensitive match found: {phc[0].name}")
                return phc[0].name

            # Bidirectional fuzzy match
            if len(transliterated) >= 3:
                phc = frappe.db.sql("""
                    SELECT name FROM `tabPHC`
                    WHERE LOWER(phc_name) LIKE CONCAT(LOWER(%(name)s), '%%')
                       OR LOWER(%(name)s) LIKE CONCAT(LOWER(phc_name), '%%')
                    ORDER BY
                        ABS(CHAR_LENGTH(phc_name) - CHAR_LENGTH(%(name)s)) ASC
                    LIMIT 1
                """, {"name": transliterated}, as_dict=True)
                if phc:
                    frappe.logger().info(
                        f"[resolve_phc] Fuzzy match: '{phc_raw}' → '{transliterated}' → '{phc[0].phc_name}'"
                    )
                    return phc[0].name

    # 4. Fallback: If no match is found, but the input is not NA, it is an "Other PHC"
    if frappe.db.exists("PHC", "Other"):
        frappe.logger().info(f"[resolve_phc] No match found for '{phc_raw}'. Falling back to 'Other'")
        return "Other"

    # If even "Other" is missing from the database
    frappe.logger().warning(f"[resolve_phc] No match found and 'Other' PHC record does not exist.")
    return None


def resolve_department(dept_raw: str, opd_category: str = None) -> str | None:
    """
    Resolve OPD department from any script to valid English option.
    Tries direct match first, then transliteration.
    """
    if not dept_raw:
        return None

    cleaned = dept_raw.strip()

    resolved = None
    # Direct match
    if cleaned in VALID_DEPARTMENTS:
        resolved = cleaned

    if not resolved:
        # Case-insensitive match
        for dept in VALID_DEPARTMENTS:
            if dept.lower() == cleaned.lower():
                resolved = dept
                break

    if not resolved:
        # Check Hindi/Marathi keyword mapping
        if cleaned in DEPARTMENT_KEYWORDS:
            resolved = DEPARTMENT_KEYWORDS[cleaned]

    if not resolved:
        # Case-insensitive keyword match
        cleaned_lower = cleaned.lower()
        for keyword, dept in DEPARTMENT_KEYWORDS.items():
            if keyword.lower() == cleaned_lower:
                resolved = dept
                break

    if not resolved:
        # Partial keyword match (input contains a known keyword)
        for keyword, dept in DEPARTMENT_KEYWORDS.items():
            if keyword in cleaned or cleaned in keyword:
                resolved = dept
                break

    if not resolved:
        # Transliterate and try again
        if is_devanagari(cleaned):
            transliterated = transliterate_to_roman(cleaned)
            if transliterated in VALID_DEPARTMENTS:
                resolved = transliterated
            else:
                for dept in VALID_DEPARTMENTS:
                    if dept.lower() == transliterated.lower():
                        resolved = dept
                        break

    # Validate against OPD category if provided
    if resolved and opd_category:
        valid_depts = []
        if opd_category == "Regular OPD":
            valid_depts = REGULAR_OPD_DEPTS
        elif opd_category == "Specialist OPD":
            valid_depts = SPECIALIST_OPD_DEPTS
        elif opd_category == "Surgical OPD":
            valid_depts = SURGICAL_OPD_DEPTS
        
        if resolved not in valid_depts:
            frappe.logger().warning(f"Resolved department '{resolved}' does not match category '{opd_category}'")
            return None

    return resolved


def resolve_opd_category(category_raw: str) -> str:
    if not category_raw:
        return ""

    cleaned = category_raw.strip().lower()
    
    # Check for Marathi/English keywords
    if "नियमित" in cleaned or "regular" in cleaned:
        return "Regular OPD"
    if "तज्ञ" in cleaned or "विशेषज्ञ" in cleaned or "specialist" in cleaned:
        return "Specialist OPD"
    if "शस्त्रक्रिया" in cleaned or "surgical" in cleaned or "मोतीबिंदू" in cleaned or "मोतियाबिंद" in cleaned or "cataract" in cleaned:
        return "Surgical OPD"

    category_map = {
        "regular": "Regular OPD",
        "regular opd": "Regular OPD",
        "specialist": "Specialist OPD",
        "specialist opd": "Specialist OPD",
        "surgical": "Surgical OPD",
        "surgical opd": "Surgical OPD",
    }
    return category_map.get(cleaned, category_raw.strip())


def resolve_referred_by_who(role_raw: str) -> str:
    if not role_raw:
        return ""
    
    cleaned = role_raw.strip().lower()
    
    # Check for keywords in Marathi/Hindi/English
    if "mmu" in cleaned:
        return "MMU Doctor"
    if "counsellor" in cleaned or "समुपदेशक" in cleaned or "परामर्शदाता" in cleaned or "mhd" in cleaned:
        return "MHD counsellor"
    if "muktipath" in cleaned or "मुक्तिपथ" in cleaned:
        return "Muktipath Karyakarta"
    if "asha" in cleaned or "आशा" in cleaned:
        return "ASHA"
    if "chw" in cleaned:
        return "CHW"
    if "supervisor" in cleaned or "सुपरवायझर" in cleaned:
        return "Supervisor"
    if "optometrist" in cleaned or "ऑप्टोमेट्रिस्ट" in cleaned:
        return "Optometrist"
    if "mpu" in cleaned or "physiotherap" in cleaned or "फिजिओथेरपिस्ट" in cleaned:
        return "MPU Physiotherapist"
        
    return role_raw.strip()


def resolve_taluka(taluka_raw: str) -> str | None:
    """
    Resolve a Glific taluka value to the Taluka doctype.
    Accepts taluka name, taluka code, or an existing document name.
    """
    if not taluka_raw:
        return None

    cleaned = taluka_raw.strip().lower()
    if not cleaned:
        return None

    # Strict map of valid Gadchiroli talukas (with English, Hindi, Marathi variants and common typos)
    taluka_map = {
        # Gadchiroli
        "gadchiroli": "Gadchiroli", "गडचिरोली": "Gadchiroli", "गड़चिरोली": "Gadchiroli",
        # Dhanora
        "dhanora": "Dhanora", "dhanura": "Dhanora", "धानोरा": "Dhanora",
        # Chamorshi
        "chamorshi": "Chamorshi", "chamorshee": "Chamorshi", "चामोर्शी": "Chamorshi",
        # Mulchera
        "mulchera": "Mulchera", "मुलचेरा": "Mulchera", "मूलचेरा": "Mulchera",
        # Desaiganj / Wadsa
        "desaiganj (wadsa)": "Desaiganj", "desaiganj(wadsa)": "Desaiganj",
        "देसाईगंज (वडसा)": "Desaiganj", "देसाईगंज(वडसा)": "Desaiganj",
        "desaiganj": "Desaiganj", "देसाईगंज": "Desaiganj", "warsa": "Desaiganj", 
        "wadsa": "Desaiganj", "वडसा": "Desaiganj", "वडसा-देसाईगंज": "Desaiganj",
        # Armori
        "armori": "Armori", "आरमोरी": "Armori",
        # Kurkheda
        "kurkheda": "Kurkheda", "कुरखेडा": "Kurkheda", "कुरखेड़ा": "Kurkheda",
        # Korchi
        "korchi": "Korchi", "कोरची": "Korchi",
        # Aheri
        "aheri": "Aheri", "अहेरी": "Aheri",
        # Sironcha
        "sironcha": "Sironcha", "सिरोंचा": "Sironcha",
        # Etapalli
        "etapalli": "Etapalli", "एटापल्ली": "Etapalli",
        # Bhamragad
        "bhamragad": "Bhamragad", "भामरागड": "Bhamragad", "भामरागढ़": "Bhamragad"
    }

    resolved = taluka_map.get(cleaned)
    if not resolved:
        # Try transliterated key lookup
        if is_devanagari(taluka_raw):
            transliterated = transliterate_to_roman(taluka_raw).strip().lower()
            resolved = taluka_map.get(transliterated)

    if resolved:
        return resolved

    # Fallback to direct DB checks (only for backward compatibility)
    if frappe.db.exists("Taluka", taluka_raw.strip()):
        return taluka_raw.strip()
    
    taluka_db = frappe.db.get_value("Taluka", {"taluka_name": taluka_raw.strip()}, "name")
    if taluka_db:
        return taluka_db

    return None


def clean_glific_value(val):
    if not val:
        return None
    val_str = str(val).strip()
    val_lower = val_str.lower()
    
    # Check for common Glific/RapidPro unresolved variable patterns
    if (
        val_lower.startswith("@contact") or 
        val_lower.startswith("contact.") or 
        val_lower.startswith("@results") or 
        val_lower.startswith("results.") or
        "contact.fields" in val_lower or
        "results." in val_lower or
        "{{" in val_str or
        "}}" in val_str
    ):
        return None
    return val_str


def parse_date(date_str):
    """Handle DD/MM/YYYY from Glific and YYYY-MM-DD from Frappe"""
    cleaned = clean_glific_value(date_str)
    if not cleaned:
        return None

    # Handle "today" / "aaj" / "आज" sent literally by Glific
    if cleaned.lower().strip() in ("today", "aaj", "आज"):
        from frappe.utils import today as frappe_today
        return getdate(frappe_today())

    # DD/MM/YYYY format from Glific
    if "/" in cleaned:
        from datetime import datetime
        try:
            return datetime.strptime(cleaned, "%d/%m/%Y").date()
        except ValueError:
            pass
            
    # DD-MM-YYYY format from Glific
    if "-" in cleaned and len(cleaned.split("-")[0]) == 2:
        from datetime import datetime
        try:
            return datetime.strptime(cleaned, "%d-%m-%Y").date()
        except ValueError:
            pass
            
    # Standard YYYY-MM-DD
    try:
        return getdate(cleaned)
    except Exception:
        return None


def parse_referral_date(date_str):
    """
    Parses the referral date received from Glific.

    Accepts:
      - 'आज' / 'aaj' / 'today' (defensive — Glific normally already
        converts the today-shortcut to an ISO date before sending)
      - ISO format (YYYY-MM-DD) — sent when Glific's "Today" branch
        computes today's date itself
      - DD/MM/YYYY — sent when the referrer typed an explicit date

    Any past date is accepted (no lower bound). Future dates are rejected.

    Raises frappe.ValidationError on invalid input. The caller must catch
    this and return it as a normal {"success": False, "error": ...} response
    rather than letting it propagate as an uncaught exception — this is what
    lets the Glific flow show the specific error text and loop back to the
    date retry node instead of landing on a generic webhook-failure message.
    """
    from datetime import datetime

    if not date_str or date_str.strip().lower() in ("आज", "aaj", "today"):
        return getdate(today())

    date_str = date_str.strip()

    # ISO format (from Glific's "Today" branch, which computes and sends
    # the date itself rather than the literal word)
    if "-" in date_str:
        try:
            return getdate(date_str)
        except Exception:
            frappe.throw("तारीख अवैध आहे. कृपया DD/MM/YYYY या स्वरूपात टाका.")

    try:
        parsed = datetime.strptime(date_str, "%d/%m/%Y").date()
    except ValueError:
        frappe.throw("तारीख अवैध आहे. कृपया DD/MM/YYYY या स्वरूपात टाका.")

    return parsed


def resolve_referrer(phone_raw: str) -> str | None:
    """
    Resolve referrer by phone number, handling various formats (with/without prefix).
    Tries: 
    1. Exact match.
    2. Normalize and match last 10 digits.
    3. Return None if no match found (avoiding random fallback).
    """
    if not phone_raw:
        return None

    # 1. Exact match on raw input (preserves underscores, suffixes, etc.)
    referrer = frappe.db.get_value("Referrer", {"phone": phone_raw}, "name")
    if referrer:
        return referrer

    # Remove all non-numeric characters
    phone_clean = "".join(filter(str.isdigit, str(phone_raw)))
    if not phone_clean:
        return None

    # 2. Exact match on cleaned digits
    referrer = frappe.db.get_value("Referrer", {"phone": phone_clean}, "name")
    if referrer:
        return referrer

    # 2. Normalize to 10 digits (last 10)
    if len(phone_clean) >= 10:
        ten_digit = phone_clean[-10:]
        # Try finding a referrer whose phone is exactly these 10 digits
        referrer = frappe.db.get_value("Referrer", {"phone": ten_digit}, "name")
        if referrer:
            return referrer
            
        # Try finding a referrer whose phone ends with these 10 digits 
        # (covers cases like '91' prefix in DB or '+91' prefix in input)
        referrer = frappe.db.get_value("Referrer", {"phone": ["like", f"%{ten_digit}"]}, "name")
        if referrer:
            return referrer

    return None


def resolve_non_visit_reason(reason_input: str) -> str | None:
    if not reason_input:
        return None
    
    reason_clean = reason_input.strip()
    
    # Mapping dictionary from code/text to the standard select option
    mapping = {
        "NV-01": "Financial Constraints",
        "Financial Constraints": "Financial Constraints",
        "पैशांची अडचण": "Financial Constraints",
        "पैसे नाही": "Financial Constraints",
        "आर्थिक अडचण": "Financial Constraints",
        "आर्थिक कारण": "Financial Constraints",
        "पैसे नसणे": "Financial Constraints",
        "आर्थिक मर्यादा": "Financial Constraints",
        "वित्तीय बाधाएं": "Financial Constraints",
        
        "NV-02": "Transport Unavailable",
        "Transport Unavailable": "Transport Unavailable",
        "वाहतूक उपलब्ध नाही": "Transport Unavailable",
        "गाडी नाही": "Transport Unavailable",
        "गाडीची सोय नाही": "Transport Unavailable",
        "वाहतूक नाही": "Transport Unavailable",
        "प्रवासाची अडचण": "Transport Unavailable",
        "परिवहन अनुपलब्ध": "Transport Unavailable",
        
        "NV-03": "Fear or Anxiety",
        "Fear or Anxiety": "Fear or Anxiety",
        "भीती वाटणे": "Fear or Anxiety",
        "भीती": "Fear or Anxiety",
        "घाबरणे": "Fear or Anxiety",
        "घाबरत आहे": "Fear or Anxiety",
        "भीती किंवा चिंता": "Fear or Anxiety",
        "डर या चिंता": "Fear or Anxiety",
        
        "NV-04": "Feeling Better",
        "Feeling Better": "Feeling Better",
        "बरे वाटत आहे": "Feeling Better",
        "बरे वाटणे": "Feeling Better",
        "तब्बेत सुधारली": "Feeling Better",
        "तब्येत सुधारली": "Feeling Better",
        "आता बरे वाटत आहे": "Feeling Better",
        "सुधारणा झाली": "Feeling Better",
        "अच्छा लगना": "Feeling Better",
        
        "NV-05": "Unaware of Appointment",
        "Unaware of Appointment": "Unaware of Appointment",
        "अपॉइंटमेंट माहित नव्हती": "Unaware of Appointment",
        "माहित नव्हते": "Unaware of Appointment",
        "माहिती नव्हती": "Unaware of Appointment",
        "तारीख माहित नव्हती": "Unaware of Appointment",
        "भेटीची माहिती नाही": "Unaware of Appointment",
        "नियुक्ति से अनभिज्ञ": "Unaware of Appointment",
        
        "NV-06": "Family Objection",
        "Family Objection": "Family Objection",
        "कुटुंबाचा विरोध": "Family Objection",
        "घरी विरोध": "Family Objection",
        "घरच्यांचा विरोध": "Family Objection",
        "घरचे नाही म्हणतात": "Family Objection",
        "कौटुंबिक आक्षेप": "Family Objection",
        "पारिवारिक आपत्ति": "Family Objection",
        
        "NV-07": "Distance Too Far",
        "Distance Too Far": "Distance Too Far",
        "खूप लांब आहे": "Distance Too Far",
        "अंतर जास्त आहे": "Distance Too Far",
        "खूप लांब": "Distance Too Far",
        "जास्त अंतर": "Distance Too Far",
        "अंतर खूप जास्त आहे": "Distance Too Far",
        "दूरी बहुत अधिक है": "Distance Too Far",
        
        "NV-08": "Other",
        "Other": "Other",
        "Other (Specify)": "Other",
        "इतर": "Other",
        "अन्य": "Other",
        "इतर (नमूद करा)": "Other",
        "अन्य (निर्दिष्ट करें)": "Other"
    }
    
    # Try direct mapping
    if reason_clean in mapping:
        return mapping[reason_clean]
        
    # Try mapping by splitting by colon (e.g. "NV-01: Financial Constraints")
    if ":" in reason_clean:
        parts = [p.strip() for p in reason_clean.split(":")]
        for part in parts:
            if part in mapping:
                return mapping[part]
                
    # If no mapping found, search case-insensitively or check substring
    for key, val in mapping.items():
        if key.lower() in reason_clean.lower():
            return val
            
    return None


def resolve_patient_health_status(status_input: str) -> str | None:
    if not status_input:
        return None
        
    clean_val = status_input.strip()
    
    mapping = {
        "Fully Recovered/Cured": "Fully Recovered/Cured",
        "Fully Recovered": "Fully Recovered/Cured",
        "Cured": "Fully Recovered/Cured",
        "पूर्ण बरे झाले": "Fully Recovered/Cured",
        "बरे झाले": "Fully Recovered/Cured",
        "पूर्णपणे बरे झाले": "Fully Recovered/Cured",
        "पूर्ण बरे": "Fully Recovered/Cured",
        "बरे": "Fully Recovered/Cured",
        "पूरी तरह से ठीक हो गया": "Fully Recovered/Cured",
        
        "Under Treatment (Ongoing)": "Under Treatment (Ongoing)",
        "Under Treatment(Ongoing)": "Under Treatment (Ongoing)",
        "Under Treatment": "Under Treatment (Ongoing)",
        "Ongoing Treatment": "Under Treatment (Ongoing)",
        "उपचार सुरू आहेत": "Under Treatment (Ongoing)",
        "उपचार चालू आहेत": "Under Treatment (Ongoing)",
        "उपचार सुरू": "Under Treatment (Ongoing)",
        "चालू उपचार": "Under Treatment (Ongoing)",
        "उपचार चालू": "Under Treatment (Ongoing)",
        "उपचार सुरू आहे": "Under Treatment (Ongoing)",
        "इलाज जारी है (जारी)": "Under Treatment (Ongoing)",
        
        "Needs Further Treatment": "Needs Further Treatment",
        "Further Treatment": "Needs Further Treatment",
        "पुढील उपचारांची गरज आहे": "Needs Further Treatment",
        "पुढील उपचार": "Needs Further Treatment",
        "आणखी उपचारांची गरज": "Needs Further Treatment",
        "पुढील गरज": "Needs Further Treatment",
        "आगे के उपचार की आवश्यकता": "Needs Further Treatment",
        
        "Condition Worsened": "Condition Worsened",
        "Worsened": "Condition Worsened",
        "तब्बेत बिघडली": "Condition Worsened",
        "तब्येत बिघडली": "Condition Worsened",
        "अवस्था बिघडली": "Condition Worsened",
        "प्रकृती बिघडली": "Condition Worsened",
        "परिस्थिति अधिकच बिघडली": "Condition Worsened",
        "स्थिति और बिगड़ गई": "Condition Worsened"
    }
    
    # Try case-insensitive matching
    for key, val in mapping.items():
        if key.lower().replace(" ", "") == clean_val.lower().replace(" ", ""):
            return val
            
    # Substring check
    for key, val in mapping.items():
        if key.lower() in clean_val.lower() or clean_val.lower() in key.lower():
            return val
            
    return mapping.get(clean_val, None)


def resolve_facility_type(facility_input: str) -> str | None:
    if not facility_input:
        return None
    facility_clean = facility_input.strip().lower()
    
    if "सर्च" in facility_clean or "search" in facility_clean:
        return "SEARCH"
    if (
        "village mh" in facility_clean or 
        "village mental health" in facility_clean or 
        "गाव" in facility_clean or 
        "व्हिलेज" in facility_clean or
        "विलेज" in facility_clean
    ):
        return "Village MH Clinic"
    if "taluka mh" in facility_clean or "taluka mental health" in facility_clean or "तालुका" in facility_clean:
        return "Taluka MH Clinic"
    if "शासकीय" in facility_clean or "सरकारी" in facility_clean or "government" in facility_clean or "सरकार" in facility_clean:
        return "Government Hospital"
    if "इतर" in facility_clean or "अन्य" in facility_clean or "other" in facility_clean:
        return "Other"
        
    facility_type_map = {
        "सर्च": "SEARCH",
        "search": "SEARCH",
        "village mh clinic": "Village MH Clinic",
        "taluka mh clinic": "Taluka MH Clinic",
        "तालुका mh क्लिनिक": "Taluka MH Clinic",
        "तालुका एमएच क्लिनिक": "Taluka MH Clinic",
        "तालुका क्लिनिक": "Taluka MH Clinic",
        "गाव mh क्लिनिक": "Village MH Clinic",
        "गावातील mh क्लिनिक": "Village MH Clinic",
        "गाव क्लिनिक": "Village MH Clinic",
        "विलेज एमएच क्लिनिक": "Village MH Clinic",
        "व्हिलेज एमएच क्लिनिक": "Village MH Clinic",
        "शासकीय": "Government Hospital",
        "शासकीय रुग्णालय": "Government Hospital",
        "सरकारी": "Government Hospital",
        "सरकारी रुग्णालय": "Government Hospital",
        "सरकारी दवाखाना": "Government Hospital",
        "government hospital": "Government Hospital",
        "government": "Government Hospital",
        "इतर": "Other",
        "अन्य": "Other",
        "other": "Other",
    }
    for key, val in facility_type_map.items():
        if key.lower() == facility_clean:
            return val
    return facility_type_map.get(facility_clean, None)


transliterate_if_devanagari = transliterate_to_roman


def send_patient_notification(patient_phone: str, patient_name: str, reference_number: str, opd_department: str) -> None:
    """
    Sends referral ID notification to the patient's WhatsApp number via Glific HSM template.
    Called internally by create_referral() after successful save.
    Fails silently — patient notification failure should NOT block referral creation.
    """
    try:
        import requests
        
        glific_api_url = frappe.conf.get("glific_api_url", "https://search.glific.com/api")
        glific_api_token = frappe.conf.get("glific_api_token")
        hsm_template_id = frappe.conf.get("patient_notification_hsm_id")
        
        if not glific_api_token or not hsm_template_id:
            frappe.logger().warning("Glific API token or HSM template ID not configured. Skipping patient notification.")
            return
        
        # Ensure phone number has country code
        phone = patient_phone.strip()
        if not phone.startswith("+") and not phone.startswith("91"):
            phone = "91" + phone
        
        headers = {
            "Authorization": glific_api_token,
            "Content-Type": "application/json",
        }
        
        # Step 1: Create or find contact in Glific by phone
        create_contact_query = """
        mutation createContact($input: ContactInput!) {
          createContact(input: $input) {
            contact { id }
          }
        }
        """
        
        # Step 2: Send HSM template message to the contact
        send_hsm_query = """
        mutation sendHsmMessage($templateId: ID!, $receiverId: ID!, $parameters: [String]!) {
          sendHsmMessage(templateId: $templateId, receiverId: $receiverId, parameters: $parameters) {
            message { id }
          }
        }
        """
        
        # Parameters for the HSM template
        parameters = [patient_name, reference_number, opd_department or "SEARCH Hospital"]
        
        # Execute create contact first
        contact_res = requests.post(
            glific_api_url,
            json={"query": create_contact_query, "variables": {"input": {"phone": phone}}},
            headers=headers,
            timeout=10
        )
        contact_res.raise_for_status()
        contact_data = contact_res.json()
        
        contact_id = None
        try:
            contact_id = contact_data["data"]["createContact"]["contact"]["id"]
        except (KeyError, TypeError):
            # If create contact failed because it exists, query contact by phone
            search_query = """
            query contact($phone: String!) {
              contact(phone: $phone) { id }
            }
            """
            search_res = requests.post(
                glific_api_url,
                json={"query": search_query, "variables": {"phone": phone}},
                headers=headers,
                timeout=10
            )
            search_res.raise_for_status()
            search_data = search_res.json()
            try:
                contact_id = search_data["data"]["contact"]["id"]
            except (KeyError, TypeError):
                frappe.logger().error(f"Could not find or create Glific contact for phone {phone}")
                return
                
        if contact_id:
            hsm_res = requests.post(
                glific_api_url,
                json={
                    "query": send_hsm_query,
                    "variables": {
                        "templateId": hsm_template_id,
                        "receiverId": contact_id,
                        "parameters": parameters
                    }
                },
                headers=headers,
                timeout=10
            )
            hsm_res.raise_for_status()
            frappe.logger().info(f"Patient notification sent to {phone} for referral {reference_number}")
        
    except Exception as e:
        frappe.logger().error(f"Failed to send patient notification: {str(e)}")


@frappe.whitelist(allow_guest=True)
def create_referral(
    contact_phone: str = "",
    referral_date_raw: str = "",
    referral_date: str = "",
    date_of_referral_raw: str = "",
    date_of_referral: str = "",
    selected_phc: str = "",
    patient_name_raw: str = "",
    father_name_raw: str = "",
    gender_raw: str = "",
    age_raw: str = "",
    village_raw: str = "",
    patient_taluka_raw: str = "",
    patient_taluka: str = "",
    taluka_raw: str = "",
    service_facility_type: str = "",
    opd_category_raw: str = "",
    opd_category: str = "",
    departments_raw: str = "",
    opd_department_raw: str = "",
    opd_department: str = "",
    other_facility_raw: str = "",
    referring_doctor_raw: str = "",
    referred_doctor_raw: str = "",
    referred_doctor: str = "",
    additional_notes_raw: str = "",
    referrer_latitude: str = "",
    referrer_longitude: str = "",
    latitude: str = "",
    longitude: str = "",
    patient_phone_raw: str = "",
    referred_by_who: str = "",
    language: str = None,
    **kwargs
) -> dict:
    try:
        # Fallback JSON body parsing if client does not send application/json Content-Type
        if not contact_phone and frappe.request:
            try:
                import json
                raw_data = frappe.request.get_data(as_text=True)
                if raw_data:
                    data = json.loads(raw_data)
                    if isinstance(data, dict):
                        contact_phone = data.get("contact_phone") or ""
                        referral_date_raw = data.get("referral_date_raw") or referral_date_raw or ""
                        referral_date = data.get("referral_date") or referral_date or ""
                        date_of_referral_raw = data.get("date_of_referral_raw") or date_of_referral_raw or ""
                        date_of_referral = data.get("date_of_referral") or date_of_referral or ""
                        selected_phc = data.get("selected_phc") or selected_phc or ""
                        patient_name_raw = data.get("patient_name_raw") or patient_name_raw or ""
                        father_name_raw = data.get("father_name_raw") or father_name_raw or ""
                        gender_raw = data.get("gender_raw") or gender_raw or ""
                        age_raw = data.get("age_raw") or age_raw or ""
                        village_raw = data.get("village_raw") or village_raw or ""
                        patient_taluka_raw = data.get("patient_taluka_raw") or patient_taluka_raw or ""
                        patient_taluka = data.get("patient_taluka") or patient_taluka or ""
                        taluka_raw = data.get("taluka_raw") or taluka_raw or ""
                        service_facility_type = data.get("service_facility_type") or service_facility_type or ""
                        opd_category_raw = data.get("opd_category_raw") or opd_category_raw or ""
                        opd_category = data.get("opd_category") or opd_category or ""
                        departments_raw = data.get("departments_raw") or departments_raw or ""
                        opd_department_raw = data.get("opd_department_raw") or opd_department_raw or ""
                        opd_department = data.get("opd_department") or opd_department or ""
                        other_facility_raw = data.get("other_facility_raw") or other_facility_raw or ""
                        referring_doctor_raw = data.get("referring_doctor_raw") or referring_doctor_raw or ""
                        referred_doctor_raw = data.get("referred_doctor_raw") or referred_doctor_raw or ""
                        referred_doctor = data.get("referred_doctor") or referred_doctor or ""
                        additional_notes_raw = data.get("additional_notes_raw") or additional_notes_raw or ""
                        referrer_latitude = data.get("referrer_latitude") or referrer_latitude or ""
                        referrer_longitude = data.get("referrer_longitude") or referrer_longitude or ""
                        latitude = data.get("latitude") or latitude or ""
                        longitude = data.get("longitude") or longitude or ""
                        patient_phone_raw = data.get("patient_phone_raw") or patient_phone_raw or ""
                        referred_by_who = data.get("referred_by_who") or referred_by_who or ""
                        language = data.get("language") or language or ""
            except Exception as e:
                frappe.log_error(f"Fallback JSON parsing failed: {str(e)}", "create_referral JSON Fallback Error")

        if not contact_phone:
            frappe.throw("contact_phone is required")

        # Clean all parameters first
        contact_phone = clean_glific_value(contact_phone)
        selected_phc = clean_glific_value(selected_phc)
        patient_name_raw = clean_glific_value(patient_name_raw)
        father_name_raw = clean_glific_value(father_name_raw)
        gender_raw = clean_glific_value(gender_raw) or ""
        age_raw = clean_glific_value(age_raw)
        village_raw = clean_glific_value(village_raw)
        patient_taluka_raw = clean_glific_value(patient_taluka_raw)
        patient_taluka = clean_glific_value(patient_taluka)
        taluka_raw = clean_glific_value(taluka_raw)
        service_facility_type = clean_glific_value(service_facility_type)
        opd_category_raw = clean_glific_value(opd_category_raw)
        opd_category = clean_glific_value(opd_category)
        departments_raw = clean_glific_value(departments_raw)
        opd_department_raw = clean_glific_value(opd_department_raw)
        opd_department = clean_glific_value(opd_department)
        other_facility_raw = clean_glific_value(other_facility_raw)
        referring_doctor_raw = clean_glific_value(referring_doctor_raw)
        referred_doctor_raw = clean_glific_value(referred_doctor_raw)
        referred_doctor = clean_glific_value(referred_doctor)
        additional_notes_raw = clean_glific_value(additional_notes_raw)
        patient_phone_raw = clean_glific_value(patient_phone_raw)
        referred_by_who = clean_glific_value(referred_by_who)
        referrer_latitude = clean_glific_value(referrer_latitude)
        referrer_longitude = clean_glific_value(referrer_longitude)
        latitude = clean_glific_value(latitude)
        longitude = clean_glific_value(longitude)

        # Defaults to SEARCH hospital if not specified for backward compatibility
        facility_type = resolve_facility_type(service_facility_type) or "SEARCH"

        valid_facilities = ["SEARCH", "Government Hospital", "Other", "Village MH Clinic", "Taluka MH Clinic"]
        if facility_type not in valid_facilities:
            frappe.throw(f"Invalid service facility type: {facility_type}")

        actual_lat = referrer_latitude or latitude
        actual_lon = referrer_longitude or longitude
        referral_date_input = (
            referral_date_raw or referral_date or date_of_referral_raw or date_of_referral
        )
        try:
            referral_date_value = parse_referral_date(referral_date_input)
        except frappe.ValidationError as e:
            # Keep HTTP status 200 so Glific's webhook routes to Success
            return {"success": False, "error": str(e)}
        patient_taluka_input = patient_taluka_raw or patient_taluka or taluka_raw
        
        # Only resolve category if facility is SEARCH
        opd_category_input = ""
        if facility_type == "SEARCH":
            opd_category_input = resolve_opd_category(opd_category_raw or opd_category)
            
        department_input = opd_department_raw or opd_department or departments_raw
        doctor_input = referring_doctor_raw or referred_doctor_raw or referred_doctor

        # Resolve referrer correctly by phone number
        referrer = resolve_referrer(contact_phone)

        # Fetch referrer details for denormalized fields
        referrer_full_name = ""
        referrer_department = ""
        if referrer:
            referrer_doc = frappe.get_doc("Referrer", referrer)
            referrer_full_name = referrer_doc.full_name or ""
            referrer_department = referrer_doc.department or ""

        # Resolve patient village, taluka and PHC using standard clean logic without dynamic creation of new profiles
        patient_village = resolve_village(village_raw)
        patient_taluka_resolved = resolve_taluka(patient_taluka_input)
        
        if not patient_taluka_resolved and patient_village:
            patient_taluka_resolved = frappe.db.get_value(
                "Village Profile", patient_village, "taluka"
            )
            
        phc = resolve_phc(selected_phc)

        # Normalize gender — supports English, Marathi, Hindi
        gender_map = {
            # English
            "male": "Male", "female": "Female", "other": "Other",
            "m": "Male", "f": "Female",
            # Marathi / Hindi
            "पुरुष": "Male", "पु": "Male",
            "स्त्री": "Female", "महिला": "Female",
            "इतर": "Other", "अन्य": "Other",
        }
        gender_clean = gender_raw.strip()
        patient_gender = gender_map.get(gender_clean.lower(), None)
        if not patient_gender:
            # Try the raw string as-is (may be Devanagari not lowerable meaningfully)
            patient_gender = gender_map.get(gender_clean, None)
        if not patient_gender:
            # Try transliterating Devanagari to English and re-matching
            if is_devanagari(gender_clean):
                gender_transliterated = transliterate_to_roman(gender_clean).lower()
                patient_gender = gender_map.get(gender_transliterated, "Other")
            else:
                patient_gender = "Other"

        # Parse age — supports "X months" (e.g. "7 months" -> 0.7 years, per spec)
        import re as _re
        age_clean = (age_raw or "").strip().lower()
        month_match = _re.search(r"(\d+(\.\d+)?)\s*(month|months|महिने|महिना|माह)", age_clean)
        try:
            if month_match:
                patient_age = round(float(month_match.group(1)) / 10, 1)
            else:
                patient_age = float(age_clean)
        except Exception:
            patient_age = 0.0

        # Resolve OPD department (SEARCH only)
        opd_dept = None
        if facility_type == "SEARCH":
            opd_dept = resolve_department(department_input, opd_category_input)

        # Transliterate names to Roman English
        patient_name = transliterate_to_roman(patient_name_raw)
        father_name = transliterate_to_roman(father_name_raw)

        # Translate additional notes to English (meaning, not transliteration)
        additional_notes = translate_to_english(additional_notes_raw)
        
        # Referring Doctor resolution
        referring_doctor = transliterate_to_roman(doctor_input)
        
        # Resolve Referred By Who role (e.g. MMU Doctor, ASHA, etc.)
        referred_by_resolved = resolve_referred_by_who(referred_by_who)

        # Save raw data exactly as received
        raw_doc = frappe.get_doc({
            "doctype": "Raw Patient Referral Data",
            "glific_contact_id": contact_phone,
            "received_at": frappe.utils.now(),
            "referral_date_raw": referral_date_raw or referral_date_input or "",
            "selected_phc": selected_phc,
            "referrer_latitude": actual_lat,
            "referrer_longitude": actual_lon,
            "patient_name_raw": patient_name_raw,
            "father_name_raw": father_name_raw,
            "gender_raw": gender_raw,
            "age_raw": age_raw,
            "village_raw": village_raw,
            "patient_taluka_raw": patient_taluka_input,
            "patient_phone_raw": patient_phone_raw,
            "service_facility_type": facility_type,
            "opd_category_raw": opd_category_raw or opd_category,
            "referred_by_who": referred_by_who,
            "departments_raw": department_input,
            "other_facility_raw": other_facility_raw or "",
            "referring_doctor_raw": doctor_input,
            "referred_doctor_raw": doctor_input,
            "additional_notes_raw": additional_notes_raw,
            "glific_referrer_name": referrer_full_name,
        })
        raw_doc.insert(ignore_permissions=True)
        frappe.db.commit()

        # Create Patient Referral
        referral_doc = frappe.get_doc({
            "doctype": "Patient Referral",
            "referral_date": referral_date_value,
            "referral_recorded_date": today(),
            "status": "Pending",
            "referrer": referrer,
            "referrer_name": referrer_full_name,
            "referrer_phone": contact_phone,
            "referrer_department": referrer_department,
            "referrer_latitude": actual_lat,
            "referrer_longitude": actual_lon,
            "referred_by_who": referred_by_resolved,
            "phc": phc or "",
            "patient_name": patient_name,
            "patient_father_name": father_name,
            "patient_gender": patient_gender,
            "patient_age": patient_age,
            "patient_village": patient_village or "",
            "patient_taluka": patient_taluka_resolved or "",
            "patient_phone": patient_phone_raw,
            "additional_notes": additional_notes,
            "service_facility_type": facility_type,
            "opd_category": opd_category_input or "",
            "opd_departments": opd_dept or "",
            "other_facility_name": other_facility_raw or "",
            "referring_doctor": referring_doctor,
            "referred_doctor": referring_doctor,
            "raw_patient_data": raw_doc.name,
        })
        referral_doc.flags.ignore_mandatory = True
        referral_doc.insert(ignore_permissions=True)
        frappe.db.commit()

        # Link raw doc back
        raw_doc.patient_referral = referral_doc.name
        raw_doc.save(ignore_permissions=True)
        frappe.db.commit()

        # Send patient notification (fails silently)
        if referral_doc.patient_phone:
            facility_desc = referral_doc.opd_departments or referral_doc.other_facility_name or referral_doc.service_facility_type
            send_patient_notification(referral_doc.patient_phone, referral_doc.patient_name, referral_doc.reference_number, facility_desc)

        # Format date as dd-mm-yyyy
        from frappe.utils import getdate
        ref_date_str = ""
        if referral_doc.referral_date:
            try:
                ref_date_str = getdate(referral_doc.referral_date).strftime("%d-%m-%Y")
            except Exception:
                ref_date_str = str(referral_doc.referral_date)

        # Clean language parameter
        lang = (clean_glific_value(language) or "mr").strip().lower()
        if lang not in ("en", "hi", "mr"):
            lang = "mr"

        # Get hospital / facility display name
        facility_display = referral_doc.service_facility_type
        patient_instruction = ""
        if referral_doc.service_facility_type == "Other" and referral_doc.other_facility_name:
            facility_display = referral_doc.other_facility_name
        elif referral_doc.service_facility_type == "SEARCH":
            facility_display = "SEARCH Hospital"
            if lang == "mr":
                patient_instruction = "कृपया SEARCH रुग्णालयातील नोंदणी खिडकीवर (Registration Desk) आपली रेफरल स्लिप दाखवा."
            elif lang == "hi":
                patient_instruction = "कृपया SEARCH अस्पताल के पंजीकरण काउंटर (Registration Desk) पर अपनी रेफरल पर्ची दिखाएं।"
            else:
                patient_instruction = "Please show your referral slip at the registration desk in SEARCH Hospital."

        # Get OPD
        opd_display = referral_doc.opd_departments or referral_doc.opd_category or "-"

        return {
            "success": True,
            "reference_number": referral_doc.reference_number,
            "patient_name": referral_doc.patient_name,
            "patient_age": referral_doc.patient_age,
            "patient_gender": referral_doc.patient_gender,
            "referral_date": ref_date_str,
            "hospital": facility_display,
            "opd": opd_display,
            "patient_instruction": patient_instruction
        }

    except frappe.ValidationError:
        raise
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "create_referral API Error")
        return {
            "success": False,
            "error": str(e)
        }


@frappe.whitelist(allow_guest=True)
def get_referral(referral_id: str = None, reference_number: str = None, supervisor_phone: str = None) -> dict:
    """
    Look up a Patient Referral by reference_number or patient_name.
    Supports guest access and returns flat and nested referral details.
    """
    try:
        ref_id = referral_id or reference_number
        if not ref_id:
            return {
                "success": False,
                "error": "Missing referral ID or reference number"
            }

        ref_id = ref_id.strip()
        referral = None
        # Try exact match on name (name in Frappe is the reference_number due to autonaming)
        # or direct search by reference_number field
        if frappe.db.exists("Patient Referral", ref_id):
            referral = frappe.get_doc("Patient Referral", ref_id)
        elif frappe.db.exists("Patient Referral", {"reference_number": ref_id}):
            referral = frappe.get_doc("Patient Referral", {"reference_number": ref_id})
        else:
            # Fuzzy match on patient_name
            name_search = transliterate_to_roman(ref_id)
            results = frappe.get_all("Patient Referral",
                filters={"patient_name": ["like", f"%{name_search}%"]},
                fields=["name"],
                order_by="referral_date desc",
                limit=1
            )
            if results:
                referral = frappe.get_doc("Patient Referral", results[0].name)

        if not referral:
            return {
                "success": False,
                "error": f"Referral {ref_id} not found"
            }

        phc_name = frappe.db.get_value("PHC", referral.phc, "phc_name") if referral.phc else ""
        patient_village_name = frappe.db.get_value(
            "Village Profile", referral.patient_village, "village_name"
        ) if referral.patient_village else ""
        referrer_name = frappe.db.get_value(
            "Referrer", referral.referrer, "full_name"
        ) if referral.referrer else ""

        # Flat structure for Glific
        res = {
            "patient_name": referral.patient_name,
            "referral_date": str(referral.referral_date),
            "opd_department": referral.opd_departments or referral.service_facility_type,
            "reference_number": referral.reference_number,
            "visit_count": referral.visit_count or 0,
            "status": referral.status,
        }

        # Nested structure for backward compatibility
        res["success"] = True
        res["referral"] = {
            "reference_number": referral.reference_number,
            "referral_date": str(referral.referral_date),
            "status": referral.status,
            "referrer_name": referrer_name,
            "referrer_phone": referral.referrer_phone,
            "phc": phc_name,
            "patient_name": referral.patient_name,
            "patient_father_name": referral.patient_father_name,
            "patient_gender": referral.patient_gender,
            "patient_age": referral.patient_age,
            "patient_village": patient_village_name,
            "patient_taluka": referral.patient_taluka or "",
            "patient_phone": referral.patient_phone or "",
            "opd_category": referral.opd_category or "",
            "opd_department": referral.opd_departments or "",
            "referred_doctor": referral.referred_doctor or "",
            "additional_notes": referral.additional_notes or "",
            "match_status": referral.match_status,
            "tribal_classification": referral.tribal_classification or "",
        }

        return res

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "get_referral API Error")
        return {
            "success": False,
            "error": str(e)
        }


@frappe.whitelist(allow_guest=True)
def record_supervisor_visit(
    referral_id: str = None,
    supervisor_phone: str = None,
    visit_date: str = None,
    patient_visited: str = None,
    facility_visited: str = None,
    confirmation_date: str = None,
    patient_health_status: str = None,
    supervisor_name: str = None,
    non_visit_reason: str = None,
    **kwargs
) -> dict:
    """
    Records a supervisor follow-up visit and updates the Patient Referral status.
    """
    try:
        # Try parsing JSON body if available (Glific uses POST)
        if frappe.request:
            try:
                import json
                raw_data = frappe.request.get_data(as_text=True)
                if raw_data:
                    data = json.loads(raw_data)
                    if isinstance(data, dict):
                        if not referral_id:
                            referral_id = data.get("referral_id") or data.get("reference_number")
                        if not supervisor_phone:
                            supervisor_phone = data.get("supervisor_phone") or data.get("contact_phone")
                        if not visit_date:
                            visit_date = data.get("visit_date") or data.get("date")
                        if not patient_visited:
                            patient_visited = data.get("patient_visited") or data.get("visited")
                        if not facility_visited:
                            facility_visited = data.get("facility_visited")
                        if not confirmation_date:
                            confirmation_date = data.get("confirmation_date")
                        if not patient_health_status:
                            patient_health_status = data.get("patient_health_status")
                        if not supervisor_name:
                            supervisor_name = data.get("supervisor_name")
                        if not non_visit_reason:
                            non_visit_reason = data.get("non_visit_reason")
            except Exception:
                pass

        # Clean all input variables first (to handle unresolved Glific variables)
        referral_id = clean_glific_value(referral_id)
        supervisor_phone = clean_glific_value(supervisor_phone)
        visit_date = clean_glific_value(visit_date)
        patient_visited = clean_glific_value(patient_visited)
        facility_visited = clean_glific_value(facility_visited)
        confirmation_date = clean_glific_value(confirmation_date)
        patient_health_status = clean_glific_value(patient_health_status)
        supervisor_name = clean_glific_value(supervisor_name)
        non_visit_reason = clean_glific_value(non_visit_reason)

        if not referral_id:
            return {
                "success": False,
                "error": "Missing referral_id"
            }
        if not supervisor_phone:
            return {
                "success": False,
                "error": "Missing supervisor_phone"
            }
        if not visit_date:
            return {
                "success": False,
                "error": "Missing visit_date"
            }
        if not patient_visited:
            return {
                "success": False,
                "error": "Missing patient_visited"
            }

        # 1. Look up referral
        referral = None
        if frappe.db.exists("Patient Referral", referral_id):
            referral = frappe.get_doc("Patient Referral", referral_id)
        elif frappe.db.exists("Patient Referral", {"reference_number": referral_id}):
            referral = frappe.get_doc("Patient Referral", {"reference_number": referral_id})

        if not referral:
            return {
                "success": False,
                "error": f"Referral {referral_id} not found"
            }

        # 2. Validate status allows new visits
        if referral.status not in ("Pending", "Follow-up In Progress"):
            return {
                "success": False,
                "error": f"Referral {referral_id} has status '{referral.status}' and cannot accept new visits"
            }

        # 3. Validate visit count
        current_count = referral.visit_count or 0
        if current_count >= 3:
            return {
                "success": False,
                "error": f"Referral {referral_id} already has 3 visits recorded. No further visits allowed."
            }

        # 4. Parse and validate dates
        from frappe.utils import getdate, today
        visit_date_parsed = parse_date(visit_date)
        if not visit_date_parsed:
            # Default to today if visit_date is missing/unresolved — supervisor records in real time
            frappe.logger().warning(
                f"[record_supervisor_visit] visit_date '{visit_date}' could not be parsed, defaulting to today"
            )
            visit_date_parsed = getdate(today())
            
        if visit_date_parsed > getdate(today()):
            return {
                "success": False,
                "error": "Visit date cannot be in the future"
            }
        if visit_date_parsed < getdate(referral.referral_date):
            return {
                "success": False,
                "error": "Visit date cannot be before referral date"
            }

        # 5. Create Supervisor Visit child record
        new_visit_number = current_count + 1
        # Support English, Marathi (होय/हो), Hindi (हाँ) affirmatives
        YES_VALUES = ("yes", "1", "true", "होय", "हो", "हाँ", "han", "hoy")
        is_visited = bool(patient_visited and patient_visited.strip().lower() in YES_VALUES)

        # Handle fields based on visited state
        if is_visited:
            confirmation_date_parsed = parse_date(confirmation_date)
            # If confirmation_date parsing failed, fallback to visit_date
            if not confirmation_date_parsed:
                confirmation_date_parsed = visit_date_parsed
                
            resolved_facility = resolve_facility_type(facility_visited)
            reason_code = None
        else:
            confirmation_date_parsed = None
            resolved_facility = None
            
            reason_code = resolve_non_visit_reason(non_visit_reason)

        referral.append("supervisor_visits", {
            "visit_number": new_visit_number,
            "visit_date": visit_date_parsed,
            "patient_visited": 1 if is_visited else 0,
            "facility_visited": resolved_facility,
            "confirmation_date": confirmation_date_parsed,
            "patient_health_status": resolve_patient_health_status(patient_health_status) if is_visited else None,
            "non_visit_reason_code": reason_code if not is_visited else None,
            "supervisor_name": supervisor_name,
            "supervisor_phone": supervisor_phone,
        })

        # 6. Update referral status (state machine)
        if is_visited:
            referral.status = "Visited"
            referral.facility_visited = resolved_facility
            referral.visit_date = confirmation_date_parsed
        elif new_visit_number >= 3:
            referral.status = "Closed - Not Visited"
        else:
            referral.status = "Follow-up In Progress"

        # 7. Update visit count
        referral.visit_count = new_visit_number

        # Auto-heal legacy referrals: populate patient_taluka from village if missing
        if not referral.patient_taluka and referral.patient_village:
            referral.patient_taluka = frappe.db.get_value("Village Profile", referral.patient_village, "taluka")

        # 8. Save
        referral.flags.ignore_mandatory = True
        referral.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "status": referral.status,
            "visit_number": new_visit_number,
        }
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "record_supervisor_visit API Error")
        return {
            "success": False,
            "error": str(e)
        }


MHD_INFO_SOURCE_MAP = {
    "patient": "Patient", "पेशंट": "Patient", "रुग्ण": "Patient",
    "relative": "Relative", "नातेवाईक": "Relative", "रिश्तेदार": "Relative",
    "neighbor": "Neighbor", "neighbour": "Neighbor", "शेजारी": "Neighbor", "पड़ोसी": "Neighbor",
    "other": "Other", "इतर": "Other", "अन्य": "Other",
}

MHD_DRINKING_PATTERN_MAP = {
    "regular": "Regular", "नियमित": "Regular",
    "binge on regular": "Binge on Regular", "नियमितपणे अति खाणे": "Binge on Regular",
    "binge": "Binge", "अति खाणे": "Binge", "बिंज": "Binge",
    "occasional": "Occasional", "अधूनमधून": "Occasional", "प्रासंगिक": "Occasional",
}

MHD_ALCOHOL_TYPE_MAP = {
    "country": "Country", "देशी": "Country",
    "imfl": "IMFL", "विदेशी (imfl)": "IMFL", "विदेशी": "IMFL",
    "moha / gul": "Moha / Gul", "moha/gul": "Moha / Gul", "मोहा / गुळ": "Moha / Gul", "मोहा / गुल": "Moha / Gul",
    "tadi / sindhi / gorga": "Tadi / Sindhi / Gorga", "ताडी / सिंधी / गोरगा": "Tadi / Sindhi / Gorga", "ताड़ी / सिंधी / गोरगा": "Tadi / Sindhi / Gorga",
    "none": "None", "काहीही नाही": "None", "कोई नहीं": "None",
}


def resolve_mhd_select(value_raw: str, mapping: dict) -> str | None:
    if not value_raw:
        return None
    return mapping.get(value_raw.strip().lower(), value_raw.strip())


@frappe.whitelist(allow_guest=True)
def record_mhd_followup(
    referral_id: str = None,
    mhd_counselor_phone: str = None,
    followup_day_offset: str = None,
    patient_info_source: str = None,
    days_drank_last_15: str = None,
    notable_incident: str = None,
    current_complaints: str = None,
    drinking_pattern: str = None,
    alcohol_type: str = None,
    quantity_ml_per_day: str = None,
    frequency_per_day: str = None,
    drank_today: str = None,
    family_opinion: str = None,
    counselor_observation: str = None,
    mhd_counselor_name: str = None,
    **kwargs
) -> dict:
    """
    Records an MHD counsellor follow-up (30/90-day addiction-management
    check-in) against a Patient Referral. Called by Glific's 'MHD Followup'
    flow, entered from the Supervisor Followup flow's MHD Followup branch.
    """
    try:
        import json
        if frappe.request:
            try:
                raw_data = frappe.request.get_data(as_text=True)
                if raw_data:
                    data = json.loads(raw_data)
                    if isinstance(data, dict):
                        referral_id = referral_id or data.get("referral_id")
                        mhd_counselor_phone = mhd_counselor_phone or data.get("mhd_counselor_phone")
                        followup_day_offset = followup_day_offset or data.get("followup_day_offset")
                        patient_info_source = patient_info_source or data.get("patient_info_source")
                        days_drank_last_15 = days_drank_last_15 or data.get("days_drank_last_15") or data.get("days_drank_last")
                        notable_incident = notable_incident or data.get("notable_incident")
                        current_complaints = current_complaints or data.get("current_complaints")
                        drinking_pattern = drinking_pattern or data.get("drinking_pattern")
                        alcohol_type = alcohol_type or data.get("alcohol_type")
                        quantity_ml_per_day = quantity_ml_per_day or data.get("quantity_ml_per_day") or data.get("quantity_ml")
                        frequency_per_day = frequency_per_day or data.get("frequency_per_day") or data.get("frequency")
                        drank_today = drank_today or data.get("drank_today")
                        family_opinion = family_opinion or data.get("family_opinion")
                        counselor_observation = counselor_observation or data.get("counselor_observation")
                        mhd_counselor_name = mhd_counselor_name or data.get("mhd_counselor_name")
            except Exception:
                pass

        referral_id = clean_glific_value(referral_id)
        mhd_counselor_phone = clean_glific_value(mhd_counselor_phone)
        mhd_counselor_name = clean_glific_value(mhd_counselor_name)
        followup_day_offset = clean_glific_value(followup_day_offset)
        patient_info_source = clean_glific_value(patient_info_source)
        days_drank_last_15 = clean_glific_value(days_drank_last_15)
        notable_incident = clean_glific_value(notable_incident)
        current_complaints = clean_glific_value(current_complaints)
        drinking_pattern = clean_glific_value(drinking_pattern)
        alcohol_type = clean_glific_value(alcohol_type)
        quantity_ml_per_day = clean_glific_value(quantity_ml_per_day)
        frequency_per_day = clean_glific_value(frequency_per_day)
        drank_today = clean_glific_value(drank_today)
        family_opinion = clean_glific_value(family_opinion)
        counselor_observation = clean_glific_value(counselor_observation)

        if not referral_id:
            return {"success": False, "error": "Missing referral_id"}
        if not mhd_counselor_phone:
            return {"success": False, "error": "Missing mhd_counselor_phone"}

        referral = None
        if frappe.db.exists("Patient Referral", referral_id):
            referral = frappe.get_doc("Patient Referral", referral_id)
        elif frappe.db.exists("Patient Referral", {"reference_number": referral_id}):
            referral = frappe.get_doc("Patient Referral", {"reference_number": referral_id})

        if not referral:
            return {"success": False, "error": f"Referral {referral_id} not found"}

        try:
            days_drank_int = int(days_drank_last_15) if days_drank_last_15 else 0
        except ValueError:
            days_drank_int = 0
        days_drank_int = max(0, min(15, days_drank_int))

        try:
            qty_ml_int = int(quantity_ml_per_day) if quantity_ml_per_day else 0
        except ValueError:
            qty_ml_int = 0

        try:
            freq_int = int(frequency_per_day) if frequency_per_day else 0
        except ValueError:
            freq_int = 0

        YES_VALUES = ("yes", "1", "true", "होय", "हो", "हाँ", "han", "hoy")
        drank_today_flag = 1 if (drank_today and drank_today.strip().lower() in YES_VALUES) else 0

        day_offset_clean = "90" if "90" in (followup_day_offset or "") else "30"

        referral.append("mhd_followups", {
            "followup_day_offset": day_offset_clean,
            "visit_date": getdate(today()),
            "patient_info_source": resolve_mhd_select(patient_info_source, MHD_INFO_SOURCE_MAP),
            "days_drank_last_15": days_drank_int,
            "notable_incident": notable_incident,
            "current_complaints": current_complaints,
            "drinking_pattern": resolve_mhd_select(drinking_pattern, MHD_DRINKING_PATTERN_MAP),
            "alcohol_type": resolve_mhd_select(alcohol_type, MHD_ALCOHOL_TYPE_MAP),
            "quantity_ml_per_day": qty_ml_int,
            "frequency_per_day": freq_int,
            "drank_today": drank_today_flag,
            "family_opinion": family_opinion,
            "counselor_observation": counselor_observation,
            "mhd_counselor_name": mhd_counselor_name,
            "mhd_counselor_phone": mhd_counselor_phone,
        })

        referral.visit_count = (referral.visit_count or 0) + 1

        referral.flags.ignore_mandatory = True
        referral.save(ignore_permissions=True)
        frappe.db.commit()

        return {"success": True, "followup_day_offset": day_offset_clean}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "record_mhd_followup API Error")
        return {"success": False, "error": str(e)}


def get_glific_contact_fields(phone: str) -> dict:
    try:
        import requests
        glific_api_url = frappe.conf.get("glific_api_url", "https://search.glific.com/api")
        glific_token = frappe.conf.get("glific_token")
        if not glific_token:
            return {}
            
        headers = {
            "Authorization": glific_token,
            "Content-Type": "application/json",
        }
        
        query = """
        query contact($phone: String!) {
          contact(phone: $phone) {
            id
            fields
          }
        }
        """
        res = requests.post(
            glific_api_url,
            json={"query": query, "variables": {"phone": phone}},
            headers=headers,
            timeout=10
        )
        res.raise_for_status()
        data = res.json()
        return data.get("data", {}).get("contact", {}).get("fields", {}) or {}
    except Exception as e:
        frappe.logger().error(f"Failed to fetch Glific contact fields: {str(e)}")
        return {}


@frappe.whitelist(allow_guest=True)
def get_pending_followups(
    supervisor_phone: str = None, 
    village_name: str = None, 
    taluka_name: str = None, 
    month: int = None, 
    year: int = None, 
    duration: str = None,
    **kwargs
) -> dict:
    """
    Returns pending follow-up referrals formatted as a ready-to-send
    Marathi WhatsApp message, optionally filtered by village name and taluka name,
    along with patients who visited SEARCH in the specified duration.

    Called by Glific 'Follow-up List' flow.
    """
    # Try parsing JSON body if available (Glific uses POST)
    if frappe.request:
        try:
            import json
            raw_data = frappe.request.get_data(as_text=True)
            if raw_data:
                data = json.loads(raw_data)
                if isinstance(data, dict):
                    if not supervisor_phone:
                        supervisor_phone = data.get("supervisor_phone")
                    if not village_name:
                        village_name = data.get("village") or data.get("village_name") or data.get("village_input")
                    if not taluka_name:
                        taluka_name = data.get("taluka") or data.get("taluka_name") or data.get("taluka_input")
                    if not month:
                        month = data.get("month")
                    if not year:
                        year = data.get("year")
                    if not duration:
                        duration = data.get("duration") or data.get("duration_months")
        except Exception:
            pass

    month = month or kwargs.get("month")
    year = year or kwargs.get("year")
    duration = duration or kwargs.get("duration") or kwargs.get("duration_months")

    supervisor_phone = clean_glific_value(supervisor_phone)
    village_name = clean_glific_value(village_name)
    taluka_name = clean_glific_value(taluka_name)

    # If village_name is not provided in body, try querying Glific contact fields
    if not village_name and supervisor_phone:
        fields = get_glific_contact_fields(supervisor_phone)
        village_name = fields.get("followup_list_village_name")
        if not taluka_name:
            taluka_name = fields.get("followup_list_taluka_name")

    # 1. Calculate duration date filters
    from frappe.utils import today, getdate, add_months
    import calendar

    now_date = getdate(today())
    start_date = None
    end_date = None
    duration_label_mr = "या महिन्यातील"

    # Clean duration input
    duration_clean = (clean_glific_value(duration) or "").strip().lower()
    
    # Map Marathi duration inputs to standard English values
    marathi_duration_map = {
        "या महिन्यातील": "this_month",
        "या महिन्यात": "this_month",
        "या": "this_month",
        "चालू महिना": "this_month",
        "चालू महिन्यातील": "this_month",
        "गेल्या महिन्यातील": "last_month",
        "गेल्या महिन्यात": "last_month",
        "मागील महिना": "last_month",
        "मागील महिन्यातील": "last_month",
        "गेल्या ३ महिन्यांतील": "last_3_months",
        "गेल्या ३ महिन्यात": "last_3_months",
        "मागील ३ महिने": "last_3_months",
        "मागील ३ महिन्यांतील": "last_3_months",
        "गेल्या ६ महिन्यांतील": "past_6_months",
        "गेल्या ६ महिन्यात": "past_6_months",
        "मागील ६ महिने": "past_6_months",
        "मागील ६ महिन्यांतील": "past_6_months",
        "सर्व काळातील": "all_time",
        "सर्व": "all_time",
        "सर्व काळ": "all_time",
        "सर्वकाळ": "all_time"
    }

    if duration_clean in marathi_duration_map:
        duration_clean = marathi_duration_map[duration_clean]
    else:
        # Fallback substring checks for extra safety
        if "सर्व" in duration_clean:
            duration_clean = "all_time"
        elif "३" in duration_clean or "3" in duration_clean:
            duration_clean = "last_3_months"
        elif "६" in duration_clean or "6" in duration_clean:
            duration_clean = "past_6_months"
        elif "गेल्या" in duration_clean or "मागील" in duration_clean or "last" in duration_clean:
            duration_clean = "last_month"
        elif "या" in duration_clean or "चालू" in duration_clean or "this" in duration_clean:
            duration_clean = "this_month"

    if duration_clean in ("this_month", "this month", "this"):
        start_date = getdate(f"{now_date.year}-{now_date.month:02d}-01")
        _, last_day = calendar.monthrange(now_date.year, now_date.month)
        end_date = getdate(f"{now_date.year}-{now_date.month:02d}-{last_day:02d}")
        duration_label_mr = "या महिन्यातील"
        
    elif duration_clean in ("last_month", "last month", "last"):
        prev_date = add_months(now_date, -1)
        start_date = getdate(f"{prev_date.year}-{prev_date.month:02d}-01")
        _, last_day = calendar.monthrange(prev_date.year, prev_date.month)
        end_date = getdate(f"{prev_date.year}-{prev_date.month:02d}-{last_day:02d}")
        duration_label_mr = "गेल्या महिन्यातील"
        
    elif duration_clean in ("last_3_months", "last 3 months", "3"):
        start_date = add_months(now_date, -3)
        end_date = now_date
        duration_label_mr = "गेल्या ३ महिन्यांतील"
        
    elif duration_clean in ("past_6_months", "past 6 months", "6"):
        start_date = add_months(now_date, -6)
        end_date = now_date
        duration_label_mr = "गेल्या ६ महिन्यांतील"
        
    elif duration_clean in ("all_time", "all time", "all"):
        start_date = None
        end_date = None
        duration_label_mr = "सर्व काळातील"
    else:
        # Fallback to month/year if specifically provided as integers
        if month:
            target_year = year or now_date.year
            target_month = month
            try:
                target_month = int(target_month)
                target_year = int(target_year)
            except (ValueError, TypeError):
                target_month = now_date.month
                target_year = now_date.year
            if 1 <= target_month <= 12:
                _, last_day = calendar.monthrange(target_year, target_month)
                start_date = getdate(f"{target_year}-{target_month:02d}-01")
                end_date = getdate(f"{target_year}-{target_month:02d}-{last_day:02d}")
                month_names_mr = {
                    1: "जानेवारी", 2: "फेब्रुवारी", 3: "मार्च", 4: "एप्रिल",
                    5: "मे", 6: "जून", 7: "जुलै", 8: "ऑगस्ट",
                    9: "सप्टेंबर", 10: "ऑक्टोबर", 11: "नोव्हेंबर", 12: "डिसेंबर"
                }
                duration_label_mr = f"{month_names_mr.get(target_month, '')} {target_year} मधील"
        else:
            # Default to this month
            start_date = getdate(f"{now_date.year}-{now_date.month:02d}-01")
            _, last_day = calendar.monthrange(now_date.year, now_date.month)
            end_date = getdate(f"{now_date.year}-{now_date.month:02d}-{last_day:02d}")
            duration_label_mr = "या महिन्यातील"

    # Set up pending referrals filters
    filters = {"status": ["in", ["Pending", "Follow-up In Progress"]]}
    
    if start_date:
        if end_date:
            filters["referral_date"] = ["between", [start_date, end_date]]
        else:
            filters["referral_date"] = [">=", start_date]
    elif end_date:
        filters["referral_date"] = ["<=", end_date]

    resolved_taluka = None
    if taluka_name:
        resolved_taluka = resolve_taluka(taluka_name)
        if not resolved_taluka:
            return {
                "formatted_text": f"❌ तालुका '{taluka_name}' आढळला नाही. कृपया तालुका तपासा आणि पुन्हा प्रयत्न करा.",
                "count": 0,
            }
        filters["patient_taluka"] = resolved_taluka

    resolved_village_name_mr = None
    village_id = None
    if village_name:
        village_id = resolve_village(village_name)
        if not village_id:
            return {
                "formatted_text": f"❌ गाव '{village_name}' आढळले नाही. कृपया गावाचे नाव तपासा आणि पुन्हा प्रयत्न करा.",
                "count": 0,
            }
        filters["patient_village"] = village_id
        resolved_village_name_mr = frappe.db.get_value("Village Profile", village_id, "village_name_marathi") or village_name

    # 1. Fetch pending referrals
    referrals = frappe.get_all(
        "Patient Referral",
        filters=filters,
        fields=[
            "reference_number",
            "patient_name",
            "patient_age",
            "patient_gender",
            "patient_village",
            "patient_taluka",
            "referral_date",
            "service_facility_type",
            "opd_departments",
            "opd_category",
            "other_facility_name",
            "visit_count",
        ],
        order_by="referral_date asc, patient_village asc",
    )

    # 2. Fetch visited SEARCH referrals
    # Build visited query conditions dynamically based on filters
    query_conditions = [
        "pr.status = 'Visited'",
        "sv.facility_visited = 'SEARCH'",
        "sv.patient_visited = 1"
    ]
    query_args = []
    
    if start_date:
        query_conditions.append("sv.visit_date >= %s")
        query_args.append(start_date)
    if end_date:
        query_conditions.append("sv.visit_date <= %s")
        query_args.append(end_date)

    if village_id:
        query_conditions.append("pr.patient_village = %s")
        query_args.append(village_id)
    elif resolved_taluka:
        query_conditions.append("pr.patient_taluka = %s")
        query_args.append(resolved_taluka)

    sql_query = f"""
        SELECT DISTINCT 
            pr.reference_number, pr.patient_name, pr.patient_age, pr.patient_gender,
            pr.patient_village, pr.patient_taluka, pr.referral_date, pr.service_facility_type,
            pr.opd_departments, pr.opd_category, pr.other_facility_name, pr.visit_count, sv.visit_date
        FROM `tabPatient Referral` pr
        JOIN `tabSupervisor Visit` sv ON sv.parent = pr.name
        WHERE {" AND ".join(query_conditions)}
        ORDER BY sv.visit_date DESC
    """
    visited_referrals = []
    if village_id or resolved_taluka:
        visited_referrals = frappe.db.sql(sql_query, tuple(query_args), as_dict=True)

    if not referrals and not visited_referrals:
        if village_name:
            return {
                "formatted_text": f"या गावात ({resolved_village_name_mr or village_name}) सध्या कोणतेही प्रलंबित किंवा भेट दिलेले रुग्ण नाहीत ✅",
                "count": 0,
            }
        return {
            "formatted_text": "सध्या कोणतेही प्रलंबित फॉलो-अप नाहीत ✅",
            "count": 0,
        }

    # Resolve village links to display names
    village_names = {}
    all_village_ids = set()
    for r in referrals:
        if r.patient_village:
            all_village_ids.add(r.patient_village)
    for r in visited_referrals:
        if r.patient_village:
            all_village_ids.add(r.patient_village)

    if all_village_ids:
        for v in frappe.get_all(
            "Village Profile",
            filters={"name": ["in", list(all_village_ids)]},
            fields=["name", "village_name", "village_name_marathi"],
        ):
            village_names[v.name] = v.village_name_marathi or v.village_name

    gender_mr = {"Male": "पुरुष", "Female": "स्त्री", "Other": "इतर"}
    dept_map_mr = {
        "Medicine": "औषध",
        "Gynaecology": "स्त्रीरोग",
        "Orthopedics": "अस्थिरोग",
        "Spine": "मणका",
        "Surgery": "शस्त्रक्रिया",
        "Dental": "दंत",
        "Psychiatry": "मानसोपचार",
        "Rheumatology": "संधिवात",
        "Cardiology": "हृदयरोग",
        "Dermatology": "त्वचारोग",
        "Diabetology": "मधुमेह",
        "ENT": "कान नाक घसा",
        "Gastrology": "पोटरोग",
        "Head & Neck": "डोके आणि मान",
        "Neurology + Epilepsy": "अपस्मार/मज्जातंतू",
        "Oncology": "कर्करोग",
        "Pulmonology": "श्वसन/फुफ्फुस",
        "Sickle Cell": "सिकल सेल",
        "Cataract Surgery": "मोतीबिंदू शस्त्रक्रिया",
        "Ophthalmology": "नेत्र",
        "Plastic Surgery": "प्लास्टिक सर्जरी",
        "Urology": "मूत्ररोग",
        "Pain Management": "वेदना व्यवस्थापन",
        "Others": "इतर",
        "Regular OPD": "नियमित ओपीडी",
        "Specialist OPD": "तज्ञ ओपीडी",
        "Surgical OPD": "शस्त्रक्रिया ओपीडी"
    }

    lines = ["*रेफर सर्च — फॉलो-अप यादी* 📋", ""]

    # Format Pending Section
    if referrals:
        # Group by referral date, then village
        grouped = defaultdict(lambda: defaultdict(list))
        for r in referrals:
            date_str = r.referral_date.strftime("%d/%m/%y")
            village_display = village_names.get(r.patient_village, r.patient_village or "गाव नोंद नाही")
            grouped[date_str][village_display].append(r)

        for date_str in sorted(
            grouped.keys(),
            key=lambda d: frappe.utils.getdate("20" + d.split("/")[2] + "-" + d.split("/")[1] + "-" + d.split("/")[0]),
        ):
            villages = grouped[date_str]
            for village, patients in villages.items():
                lines.append(f"🟦 *{date_str}* ({village})")
                for i, p in enumerate(patients, 1):
                    if p.service_facility_type == "Other" and p.other_facility_name:
                        hospital = p.other_facility_name
                    else:
                        hospital = p.service_facility_type or "SEARCH"

                    dept = p.opd_departments or p.opd_category or "-"
                    dept_mr = dept_map_mr.get(dept, dept)
                    gender = gender_mr.get(p.patient_gender, p.patient_gender or "")

                    lines.append(f"{i}) *{p.patient_name}*")
                    lines.append(f"वय-{p.patient_age}/{gender}")
                    lines.append(f"🏥 {hospital} | विभाग: {dept_mr}")
                    lines.append(f"🔖 {p.reference_number}")
                    if p.visit_count and p.visit_count > 0:
                        lines.append(f"पुढील भेट क्र. {p.visit_count + 1}")
                    lines.append("──────────────────")
    else:
        if visited_referrals:
            lines.append("प्रलंबित फॉलो-अप: काहीही नाही ✅")
            lines.append("──────────────────")

    # Format Visited Section
    if visited_referrals:
        lines.append(f"✅ *{duration_label_mr} SEARCH ला भेट दिलेले रुग्ण:*")
        lines.append("")

        for i, p in enumerate(visited_referrals, 1):
            v_date_str = p.visit_date.strftime("%d/%m/%y") if p.visit_date else "-"
            village_display = village_names.get(p.patient_village, p.patient_village or "गाव नोंद नाही")
            gender = gender_mr.get(p.patient_gender, p.patient_gender or "")

            lines.append(f"{i}) *{p.patient_name}* ({village_display})")
            lines.append(f"वय-{p.patient_age}/{gender}")
            lines.append(f"📅 भेट तारीख: {v_date_str}")
            lines.append(f"🔖 {p.reference_number}")
            lines.append("──────────────────")

    formatted = "\n".join(lines).strip()

    # WhatsApp message hard limit is 4096 chars. Truncate safely if huge.
    if len(formatted) > 3900:
        formatted = formatted[:3900] + "\n\n... यादी खूप मोठी आहे. कृपया प्रशासकाशी संपर्क साधा."

    return {"formatted_text": formatted, "count": len(referrals)}


@frappe.whitelist(allow_guest=False)
def update_registration(
    reference_number: str,
    hospital_registration_number: str = "",
    visit_date: str = ""
) -> dict:
    try:
        if not frappe.db.exists("Patient Referral", reference_number):
            return {
                "success": False,
                "error": f"Referral {reference_number} not found"
            }

        doc = frappe.get_doc("Patient Referral", reference_number)
        doc.status = "Visited"
        doc.hospital_registration_number = hospital_registration_number
        doc.visit_date = visit_date or frappe.utils.today()
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "message": f"Referral {reference_number} updated to Visited"
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "update_registration API Error")
        return {
            "success": False,
            "error": str(e)
        }


@frappe.whitelist()
def translate_villages_to_marathi():
    """
    Translate all Village Profile names from English to Marathi using Google Translate.
    Requires System Manager role.
    """
    frappe.only_for("System Manager")
    
    print("\n" + "="*70)
    print("TRANSLATING VILLAGES TO MARATHI")
    print("="*70 + "\n")
    
    try:
        # Get all villages
        villages = frappe.db.get_list("Village Profile", fields=["name", "village_name"])
        total = len(villages)
        
        if total == 0:
            print("No villages found in the database.")
            return {"success": True, "message": "No villages found"}
        
        print(f"Found {total} villages to translate.\n")
        
        translated = 0
        failed = 0
        skipped = 0
        results = []
        
        for idx, v in enumerate(villages, 1):
            try:
                village_name = v.get("village_name")
                existing_marathi = frappe.db.get_value(
                    "Village Profile", 
                    v.get("name"), 
                    "village_name_marathi"
                )
                
                # Skip if already has Marathi translation
                if existing_marathi:
                    msg = f"[{idx}/{total}] {village_name}: SKIPPED (already has Marathi)"
                    print(msg)
                    results.append(msg)
                    skipped += 1
                    continue
                
                # Translate to Marathi
                try:
                    marathi = GoogleTranslator(source="en", target="mr").translate(village_name)
                except Exception as trans_err:
                    msg = f"[{idx}/{total}] {village_name}: TRANSLATION ERROR - {str(trans_err)}"
                    print(msg)
                    results.append(msg)
                    failed += 1
                    continue
                
                if marathi and marathi != village_name:
                    # Update the village
                    frappe.db.set_value(
                        "Village Profile", 
                        v.get("name"), 
                        "village_name_marathi", 
                        marathi
                    )
                    msg = f"[{idx}/{total}] {village_name:30} → {marathi}"
                    print(msg)
                    results.append(msg)
                    translated += 1
                else:
                    msg = f"[{idx}/{total}] {village_name}: FAILED (translation same as input)"
                    print(msg)
                    results.append(msg)
                    failed += 1
                    
            except Exception as e:
                msg = f"[{idx}/{total}] {village_name}: ERROR - {str(e)}"
                print(msg)
                results.append(msg)
                failed += 1
        
        # Commit all changes
        frappe.db.commit()
        
        summary = f"\n{'='*70}\nTRANSLATION COMPLETE\n{'='*70}\n  Translated: {translated}\n  Skipped: {skipped}\n  Failed: {failed}\n  Total: {total}\n"
        print(summary)
        results.append(summary)
        
        return {
            "success": True,
            "translated": translated,
            "skipped": skipped,
            "failed": failed,
            "total": total,
            "results": results
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Village Translation Error")
        return {
            "success": False,
            "error": str(e)
        }


@frappe.whitelist()
def add_phcs_with_marathi():
    """
    Add new PHCs (Murumgaon, Rangi, Karwafa, Pendhri, Godalwahi, Other) with Marathi names.
    Requires System Manager role.
    """
    frappe.only_for("System Manager")
    
    print("\n" + "="*70)
    print("ADDING NEW PHCs")
    print("="*70 + "\n")
    
    phcs_data = [
        {
            "phc_name": "Murumgaon",
            "phc_name_marathi": "मुरुमगाव",
            "code": "PHC_001"
        },
        {
            "phc_name": "Rangi",
            "phc_name_marathi": "रंगी",
            "code": "PHC_002"
        },
        {
            "phc_name": "Karwafa",
            "phc_name_marathi": "करवाफा",
            "code": "PHC_003"
        },
        {
            "phc_name": "Pendhri",
            "phc_name_marathi": "पेंढरी",
            "code": "PHC_004"
        },
        {
            "phc_name": "Godalwahi",
            "phc_name_marathi": "गोदलवाही",
            "code": "PHC_005"
        },
        {
            "phc_name": "Other",
            "phc_name_marathi": "इतर",
            "code": "PHC_006"
        }
    ]
    
    results = []
    added = 0
    skipped = 0
    failed = 0
    
    try:
        for idx, phc_data in enumerate(phcs_data, 1):
            try:
                # Check if PHC already exists
                existing = frappe.db.get_value("PHC", {"phc_name": phc_data["phc_name"]})
                if existing:
                    msg = f"[{idx}/{len(phcs_data)}] {phc_data['phc_name']}: SKIPPED (already exists)"
                    print(msg)
                    results.append(msg)
                    skipped += 1
                    continue
                
                # Create new PHC  
                phc_doc = frappe.get_doc({
                    "doctype": "PHC",
                    "phc_name": phc_data["phc_name"],
                    "phc_name_marathi": phc_data["phc_name_marathi"],
                    "code": phc_data["code"],
                    "state": "Maharashtra",
                    "district": "Gadchiroli"
                })
                phc_doc.insert(ignore_permissions=True)
                msg = f"[{idx}/{len(phcs_data)}] {phc_data['phc_name']:20} → {phc_data['phc_name_marathi']:15} ✓"
                print(msg)
                results.append(msg)
                added += 1
                
            except Exception as e:
                msg = f"[{idx}/{len(phcs_data)}] {phc_data['phc_name']}: ERROR - {str(e)}"
                print(msg)
                results.append(msg)
                failed += 1
        
        frappe.db.commit()
        
        summary = f"\n{'='*70}\nPHCS ADDED SUCCESSFULLY\n{'='*70}\n  Added: {added}\n  Skipped: {skipped}\n  Failed: {failed}\n  Total: {len(phcs_data)}\n"
        print(summary)
        results.append(summary)
        
        return {
            "success": True,
            "added": added,
            "skipped": skipped,
            "failed": failed,
            "total": len(phcs_data),
            "results": results
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Add PHCs Error")
        return {
            "success": False,
            "error": str(e)
        }


@frappe.whitelist()
def insert_samples():
    print("Clearing existing referrals...")
    frappe.db.delete("Patient Referral")
    
    samples = [
        # Referral 1: Oldest date
        {
            "contact_phone": "9876543210",
            "patient_name_raw": "Ramesh Madavi",
            "father_name_raw": "Laxman",
            "gender_raw": "Male",
            "age_raw": "48",
            "village_raw": "Alaknar",
            "patient_taluka_raw": "Dhanora",
            "service_facility_type": "SEARCH",
            "opd_category_raw": "Regular OPD",
            "departments_raw": "Medicine",
            "referral_date_raw": "01/07/2026",
            "patient_phone_raw": "9100000001",
            "referring_doctor_raw": "Dr. Patil"
        },
        # Referral 2: Government facility
        {
            "contact_phone": "9876543210",
            "patient_name_raw": "Sunita Pudo",
            "father_name_raw": "Raju",
            "gender_raw": "Female",
            "age_raw": "35",
            "village_raw": "Ambezari",
            "patient_taluka_raw": "Dhanora",
            "service_facility_type": "Government Hospital",
            "referral_date_raw": "02/07/2026",
            "patient_phone_raw": "9100000002",
            "referring_doctor_raw": "Dr. Patil"
        },
        # Referral 3: Same date, village Arjuni (Patient A)
        {
            "contact_phone": "9876543210",
            "patient_name_raw": "Vilas Atram",
            "father_name_raw": "Sukhdeo",
            "gender_raw": "Male",
            "age_raw": "50",
            "village_raw": "Arjuni",
            "patient_taluka_raw": "Dhanora",
            "service_facility_type": "Other",
            "other_facility_raw": "Civil Hospital Nagpur",
            "referral_date_raw": "05/07/2026",
            "patient_phone_raw": "9100000003",
            "referring_doctor_raw": "Dr. Patil"
        },
        # Referral 4: Same date, village Arjuni (Patient B)
        {
            "contact_phone": "9876543210",
            "patient_name_raw": "Kamla Halami",
            "father_name_raw": "Devaji",
            "gender_raw": "Female",
            "age_raw": "60",
            "village_raw": "Arjuni",
            "patient_taluka_raw": "Dhanora",
            "service_facility_type": "SEARCH",
            "opd_category_raw": "Surgical OPD",
            "departments_raw": "Ophthalmology",
            "referral_date_raw": "05/07/2026",
            "patient_phone_raw": "9100000004",
            "referring_doctor_raw": "Dr. Patil"
        },
        # Referral 5: Follow-up In Progress
        {
            "contact_phone": "9876543210",
            "patient_name_raw": "Devidas Usendi",
            "father_name_raw": "Kavdu",
            "gender_raw": "Male",
            "age_raw": "28",
            "village_raw": "Aswalpar",
            "patient_taluka_raw": "Dhanora",
            "service_facility_type": "SEARCH",
            "opd_category_raw": "Specialist OPD",
            "departments_raw": "Diabetology",
            "referral_date_raw": "07/07/2026",
            "patient_phone_raw": "9100000005",
            "referring_doctor_raw": "Dr. Patil"
        }
    ]
    
    for i, s in enumerate(samples, 1):
        res = create_referral(**s)
        if res.get("success"):
            ref_num = res.get("reference_number")
            print(f"Created referral {i}: {ref_num} for {s['patient_name_raw']}")
            
            # Record a visit for the last patient to set it to Follow-up In Progress
            if s["patient_name_raw"] == "Devidas Usendi":
                visit_res = record_supervisor_visit(
                    referral_id=ref_num,
                    supervisor_phone="9999999999",
                    visit_date="08-07-2026",
                    patient_visited="No",
                    non_visit_reason="NV-01: Financial Constraints"
                )
                if visit_res.get("success"):
                    print(f"Recorded follow-up visit for Devidas Usendi. Status: {visit_res.get('status')}")
        else:
            print(f"Failed to create referral {i}: {res.get('error')}")
            
    frappe.db.commit()
    return {"success": True}


# @frappe.whitelist()
def import_village_list() -> dict:
    """
    One-off import function to read 1500 villages from /Users/sakshi/Downloads/12 Taluka Village List (1500).xlsx
    and automatically translate village names into Marathi via deep_translator.
    """
    import openpyxl
    from deep_translator import GoogleTranslator

    file_path = "/Users/sakshi/Downloads/12 Taluka Village List (1500).xlsx"
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        return {"success": False, "error": f"Failed to open workbook: {str(e)}"}

    if "1500 Village" not in wb.sheetnames:
        return {"success": False, "error": "Sheet '1500 Village' not found in workbook"}

    ws = wb["1500 Village"]
    rows = list(ws.iter_rows(values_only=True))

    count = 0
    skipped = 0
    translated = 0
    errors = 0

    # Get max village number to increment from
    max_num_list = frappe.get_all("Village Profile", fields=["village_number"], order_by="village_number desc", limit=1)
    current_num = max_num_list[0].village_number if max_num_list else 0
    if not isinstance(current_num, int):
        current_num = 0

    translator = GoogleTranslator(source="en", target="mr")

    for idx, r in enumerate(rows[2:], start=3):
        taluka_input, village_input = r[0], r[1]
        if not village_input:
            continue

        village_name = str(village_input).strip()
        taluka_clean = str(taluka_input).strip() if taluka_input else ""

        # Check if village already exists
        if frappe.db.exists("Village Profile", village_name):
            skipped += 1
            continue

        resolved_taluka = resolve_taluka(taluka_clean)
        if resolved_taluka and not frappe.db.exists("Taluka", resolved_taluka):
            try:
                t_doc = frappe.get_doc({
                    "doctype": "Taluka",
                    "taluka_name": resolved_taluka,
                    "taluka_code": resolved_taluka[:3].upper(),
                    "district": "Gadchiroli",
                    "state": "Maharashtra"
                })
                t_doc.insert(ignore_permissions=True)
            except Exception:
                pass

        # Translate to Marathi
        village_name_marathi = None
        try:
            marathi = translator.translate(village_name)
            if marathi and marathi != village_name:
                village_name_marathi = marathi
                translated += 1
        except Exception:
            errors += 1

        current_num += 1

        doc = frappe.get_doc({
            "doctype": "Village Profile",
            "village_name": village_name,
            "village_number": current_num,
            "taluka": resolved_taluka,
            "village_name_marathi": village_name_marathi
        })
        doc.insert(ignore_permissions=True)
        count += 1

        if count % 100 == 0:
            frappe.db.commit()

    frappe.db.commit()
    return {
        "success": True,
        "imported": count,
        "skipped": skipped,
        "translated": translated,
        "translation_errors": errors
    }


@frappe.whitelist(allow_guest=True)
def search_and_resolve_village(village_input: str = None, taluka: str = None, contact_phone: str = None, language: str = None, **kwargs) -> dict:
    import json

    # Glific POSTs JSON — fall back to parsing the raw body for keys that
    # don't match the function's parameter names exactly (e.g. the flow
    # sends "taluka_input" instead of "taluka").
    if frappe.request:
        try:
            raw_data = frappe.request.get_data(as_text=True)
            if raw_data:
                data = json.loads(raw_data)
                if isinstance(data, dict):
                    village_input = village_input or data.get("village_input")
                    taluka = taluka or data.get("taluka") or data.get("taluka_input")
                    contact_phone = contact_phone or data.get("contact_phone") or data.get("phone")
                    language = language or data.get("language")
        except Exception:
            pass

    village_clean = clean_glific_value(village_input)
    taluka_clean = clean_glific_value(taluka)
    contact_phone = clean_glific_value(contact_phone)
    lang = (clean_glific_value(language) or "en").strip().lower()
    if lang not in ("en", "hi", "mr"):
        lang = "en"

    if not village_clean:
        return {
            "success": False,
            "resolved": False,
            "village_name": None,
            "formatted_text": village_msg("village_empty", lang),
            "matches": []
        }

    # 1. Try resolving exact/transliterated match using resolve_village
    resolved_village = resolve_village(village_clean)
    if resolved_village:
        v_taluka = frappe.db.get_value("Village Profile", resolved_village, "taluka")
        resolved_taluka = resolve_taluka(taluka_clean) if taluka_clean else None

        if not resolved_taluka or v_taluka == resolved_taluka:
            display_name = village_display_name(resolved_village, lang)
            return {
                "success": True,
                "resolved": True,
                "village_name": resolved_village,
                "formatted_text": village_msg("resolved", lang, name=display_name),
                "matches": []
            }

    # 2. If not resolved exactly, find similar villages under the specified taluka
    resolved_taluka = resolve_taluka(taluka_clean) if taluka_clean else None
    filters = {}
    if resolved_taluka:
        filters["taluka"] = resolved_taluka

    villages = frappe.get_all(
        "Village Profile",
        filters=filters,
        fields=["name", "village_name", "village_name_marathi"]
    )

    query = village_clean.strip().lower()
    matches = []

    for v in villages:
        name_eng = (v.village_name or "").strip().lower()
        name_mr = (v.village_name_marathi or "").strip().lower()

        score = 0
        if query == name_eng or query == name_mr:
            score = 100
        elif name_eng.startswith(query) or name_mr.startswith(query):
            score = 80
        elif query in name_eng or name_eng in query:
            score = 50
        elif query in name_mr or name_mr in query:
            score = 50
        else:
            if is_devanagari(village_clean):
                query_roman = transliterate_to_roman(village_clean).strip().lower()
                if query_roman == name_eng or query_roman.startswith(name_eng) or name_eng.startswith(query_roman):
                    score = 70
                elif query_roman in name_eng or name_eng in query_roman:
                    score = 40
            else:
                if v.village_name_marathi:
                    v_mar_roman = transliterate_to_roman(v.village_name_marathi).strip().lower()
                    if query == v_mar_roman or query.startswith(v_mar_roman) or v_mar_roman.startswith(query):
                        score = 70
                    elif query in v_mar_roman or v_mar_roman in query:
                        score = 40
        if score > 0:
            matches.append((score, v.name))

    matches.sort(key=lambda x: x[0], reverse=True)

    unique_matches = []
    seen = set()
    for score, m_name in matches:
        if m_name not in seen:
            seen.add(m_name)
            unique_matches.append(m_name)
            if len(unique_matches) >= 5:
                break

    if not unique_matches:
        return {
            "success": True,
            "resolved": False,
            "village_name": None,
            "formatted_text": village_msg("no_match_in_taluka", lang),
            "matches": []
        }

    # Persist the candidate list server-side, keyed by the WhatsApp phone
    # number, so resolve_village_selection can retrieve it later without
    # needing the flow to hand the array back.
    if contact_phone:
        frappe.cache().set_value(
            f"village_matches:{contact_phone}",
            unique_matches,
            expires_in_sec=VILLAGE_MATCH_CACHE_TTL_SECONDS,
        )
    else:
        frappe.logger().warning(
            "[search_and_resolve_village] No contact_phone provided — cannot cache matches for later selection"
        )

    lines = [village_msg("did_you_mean", lang)]
    for i, m_name in enumerate(unique_matches):
        lines.append(f"{i + 1}. {village_display_name(m_name, lang)}")
    lines.append(f"{len(unique_matches) + 1}. {village_msg('none_of_these', lang)}")

    return {
        "success": True,
        "resolved": False,
        "village_name": None,
        "formatted_text": "\n".join(lines),
        "matches": unique_matches,
    }


@frappe.whitelist(allow_guest=True)
def resolve_village_selection(selection_input: str = None, contact_phone: str = None, language: str = None, **kwargs) -> dict:
    import json

    if frappe.request:
        try:
            raw_data = frappe.request.get_data(as_text=True)
            if raw_data:
                data = json.loads(raw_data)
                if isinstance(data, dict):
                    selection_input = selection_input or data.get("selection_input")
                    contact_phone = contact_phone or data.get("contact_phone") or data.get("phone")
                    language = language or data.get("language")
        except Exception:
            pass

    sel_clean = clean_glific_value(selection_input)
    contact_phone = clean_glific_value(contact_phone)
    lang = (clean_glific_value(language) or "en").strip().lower()
    if lang not in ("en", "hi", "mr"):
        lang = "en"

    if not sel_clean:
        return {"success": False, "resolved": False, "village_name": None}
    if not contact_phone:
        return {"success": False, "resolved": False, "village_name": None, "error": "Missing contact_phone"}

    num_map = {
        "१": 1, "२": 2, "३": 3, "४": 4, "५": 5, "६": 6, "७": 7, "८": 8, "९": 9, "०": 0
    }
    sel_str = sel_clean.strip()
    for dev_digit, eng_digit in num_map.items():
        sel_str = sel_str.replace(dev_digit, str(eng_digit))

    try:
        index = int(sel_str)
    except ValueError:
        return {"success": False, "resolved": False, "village_name": None}

    matches_list = frappe.cache().get_value(f"village_matches:{contact_phone}")
    if not matches_list:
        return {
            "success": False,
            "resolved": False,
            "village_name": None,
            "error": village_msg("session_expired", lang),
        }

    if index == len(matches_list) + 1:
        return {
            "success": True,
            "resolved": False,
            "village_name": None
        }

    if 1 <= index <= len(matches_list):
        selected_name = matches_list[index - 1]
        # Clear the cache entry once resolved
        frappe.cache().delete_value(f"village_matches:{contact_phone}")
        return {
            "success": True,
            "resolved": True,
            "village_name": selected_name
        }

    return {"success": False, "resolved": False, "village_name": None}


@frappe.whitelist()
def export_translated_villages():
    import json
    villages = frappe.get_all(
        "Village Profile",
        fields=["village_name", "village_name_marathi", "taluka", "village_number"]
    )
    with open("translated_villages.json", "w", encoding="utf-8") as f:
        json.dump(villages, f, ensure_ascii=False, indent=4)
    return {"success": True, "count": len(villages)}


@frappe.whitelist()
def import_translated_villages():
    import json
    import os
    
    file_path = "translated_villages.json"
    if not os.path.exists(file_path):
        return {"success": False, "error": f"{file_path} not found"}
        
    with open(file_path, "r", encoding="utf-8") as f:
        villages = json.load(f)
        
    count = 0
    skipped = 0
    
    for v in villages:
        village_name = v.get("village_name")
        if not village_name:
            continue
            
        if frappe.db.exists("Village Profile", village_name):
            skipped += 1
            continue
            
        taluka = v.get("taluka")
        if taluka and not frappe.db.exists("Taluka", taluka):
            try:
                t_doc = frappe.get_doc({
                    "doctype": "Taluka",
                    "taluka_name": taluka,
                    "taluka_code": taluka[:3].upper(),
                    "district": "Gadchiroli",
                    "state": "Maharashtra"
                })
                t_doc.insert(ignore_permissions=True)
            except Exception:
                pass
                
        doc = frappe.get_doc({
            "doctype": "Village Profile",
            "village_name": village_name,
            "village_number": v.get("village_number"),
            "taluka": taluka,
            "village_name_marathi": v.get("village_name_marathi")
        })
        doc.insert(ignore_permissions=True)
        count += 1
        
        if count % 100 == 0:
            frappe.db.commit()
            
    frappe.db.commit()
    return {"success": True, "imported": count, "skipped": skipped}


